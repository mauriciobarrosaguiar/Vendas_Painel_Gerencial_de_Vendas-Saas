from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Mapping

import pandas as pd

from .tratamento import (
    COLUNAS_ACOES,
    COLUNAS_BUSSOLA,
    COLUNAS_CONTATO,
    COLUNAS_PAINEL,
    COLUNAS_PRODUTOS_MIX,
    TIPO_SEM_CLASSIFICACAO,
    deduplicar_exportacao_bussola,
    normalizar_cnpj,
    normalizar_ean,
    padronizar_colunas,
    preparar_acoes,
    preparar_base_vendas,
    preparar_painel_equipe,
    preparar_produtos_mix,
    renomear_alias,
    validar_colunas_esperadas,
)


ABAS_PADRAO = {
    "bussola": "Pedidos",
    "painel": "Planilha1",
    "acoes": 0,
    "produtos_mix": 0,
    "mercado_farma": 0,
    "produtos_mercado_farma": 0,
    "bussola_historico": "Pedidos",
}

ARQUIVOS_PADRAO = {
    "bussola": "data/bussola.xlsx",
    "painel": "data/PAINEL EQUIPE NORTE.xlsx",
    "acoes": "data/template_acoes_promocionais.xlsx",
    "produtos_mix": "data/template_produtos_mix.xlsx",
    "mercado_farma": "data/mercado_farma.xlsx",
    "produtos_mercado_farma": "data/produtos.xlsx",
    "bussola_historico": "data/bussola_historico.xlsx",
}


@dataclass(frozen=True)
class BasesRaw:
    bussola: pd.DataFrame
    painel: pd.DataFrame
    acoes: pd.DataFrame
    produtos_mix: pd.DataFrame
    mercado_farma: pd.DataFrame
    produtos_mercado_farma: pd.DataFrame
    bussola_historico: pd.DataFrame


def _celula_vazia(valor: object) -> bool:
    if valor is None:
        return True
    if pd.isna(valor):
        return True
    return str(valor).strip().lower() in {"", "nan", "none", "<na>", "nat", "null"}


def _colunas_unicas(cabecalho: list[object]) -> list[str]:
    nomes: list[str] = []
    contagem: dict[str, int] = {}
    for idx, valor in enumerate(cabecalho):
        nome = str(valor).strip() if not _celula_vazia(valor) else f"Unnamed: {idx}"
        ocorrencias = contagem.get(nome, 0)
        contagem[nome] = ocorrencias + 1
        nomes.append(nome if ocorrencias == 0 else f"{nome}.{ocorrencias}")
    return nomes


def ler_excel_bytes(conteudo: bytes, sheet_name: str | int = 0) -> pd.DataFrame:
    if not conteudo:
        return pd.DataFrame()
    return pd.read_excel(BytesIO(conteudo), sheet_name=sheet_name, dtype=str, engine="openpyxl")


def ler_painel_bytes(conteudo: bytes) -> pd.DataFrame:
    if not conteudo:
        return pd.DataFrame()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb[ABAS_PADRAO["painel"]] if ABAS_PADRAO["painel"] in wb.sheetnames else wb.worksheets[0]
        cabecalho: list[str] | None = None
        linhas: list[list[object]] = []
        vazias_consecutivas = 0
        for linha in ws.iter_rows(values_only=True):
            valores = list(linha)
            if all(_celula_vazia(valor) for valor in valores):
                if cabecalho is not None:
                    vazias_consecutivas += 1
                    if vazias_consecutivas >= 500:
                        break
                continue
            vazias_consecutivas = 0
            if cabecalho is None:
                cabecalho = _colunas_unicas(valores)
                continue
            if len(valores) < len(cabecalho):
                valores.extend([""] * (len(cabecalho) - len(valores)))
            linhas.append(valores[: len(cabecalho)])
        wb.close()
        return pd.DataFrame(linhas, columns=cabecalho or [])
    except Exception:
        try:
            return ler_excel_bytes(conteudo, ABAS_PADRAO["painel"])
        except ValueError:
            return ler_excel_bytes(conteudo, 0)


def ler_csv_bytes(conteudo: bytes) -> pd.DataFrame:
    if not conteudo:
        return pd.DataFrame()
    return pd.read_csv(BytesIO(conteudo), dtype=str, sep=None, engine="python")


def ler_base_bytes(chave: str, conteudo: bytes, nome_arquivo: str | None = None) -> pd.DataFrame:
    if nome_arquivo and nome_arquivo.lower().endswith(".csv"):
        return ler_csv_bytes(conteudo)
    if chave == "painel":
        return ler_painel_bytes(conteudo)
    return ler_excel_bytes(conteudo, ABAS_PADRAO.get(chave, 0))


def carregar_bases_de_bytes(arquivos: Mapping[str, bytes]) -> BasesRaw:
    return BasesRaw(
        bussola=ler_base_bytes("bussola", arquivos.get("bussola", b"")),
        painel=ler_base_bytes("painel", arquivos.get("painel", b"")),
        acoes=ler_base_bytes("acoes", arquivos.get("acoes", b"")),
        produtos_mix=ler_base_bytes("produtos_mix", arquivos.get("produtos_mix", b"")),
        mercado_farma=ler_base_bytes("mercado_farma", arquivos.get("mercado_farma", b"")),
        produtos_mercado_farma=ler_base_bytes("produtos_mercado_farma", arquivos.get("produtos_mercado_farma", b"")),
        bussola_historico=ler_base_bytes("bussola_historico", arquivos.get("bussola_historico", b"")),
    )


def carregar_bases_de_arquivos(root_dir: Path | str) -> BasesRaw:
    root = Path(root_dir)
    dados: dict[str, bytes] = {}
    for chave, rel in ARQUIVOS_PADRAO.items():
        path = root / rel
        dados[chave] = path.read_bytes() if path.exists() else b""
    return carregar_bases_de_bytes(dados)


def validar_upload_generico(chave: str, conteudo: bytes, nome_arquivo: str | None = None) -> tuple[bool, str]:
    try:
        bruto = ler_base_bytes(chave, conteudo, nome_arquivo)
    except Exception as exc:
        return False, f"Nao consegui ler o arquivo enviado: {exc}"
    if bruto.empty:
        return False, "O arquivo enviado esta vazio."

    base = padronizar_colunas(bruto)
    if chave in {"bussola", "bussola_historico"}:
        aliases = {
            "cnpj_pdv": ["CNPJ PDV", "CNPJ", "CNPJ CLIENTE"],
            "ean": ["EAN", "CODIGO DE BARRAS", "COD BARRAS"],
            "produto": ["PRODUTO", "DESCRICAO", "NOME PRODUTO"],
            "quantidade_solicitada": ["QUANTIDADE SOLICITADA", "QTD SOLICITADA", "QTD"],
            "quantidade_atendida": ["QUANTIDADE ATENDIDA", "QTD ATENDIDA"],
            "quantidade_faturada": ["QUANTIDADE FATURADA", "QTD FATURADA"],
        }
        for destino, nomes in aliases.items():
            base = renomear_alias(base, destino, nomes)
        minimas = ["cnpj_pdv", "ean", "produto"]
        faltantes = [coluna for coluna in minimas if coluna not in base.columns]
        if faltantes:
            return False, "A base Bussola precisa conter ao menos CNPJ PDV, EAN e Produto."
        if not any(coluna in base.columns for coluna in ["quantidade_solicitada", "quantidade_atendida", "quantidade_faturada"]):
            return False, "A base Bussola precisa conter quantidade_solicitada, quantidade_atendida ou quantidade_faturada."
    elif chave == "painel":
        tratado = preparar_painel_equipe(bruto)
        cnpjs_validos = int(tratado["cnpj_limpo"].dropna().astype(str).str.strip().ne("").sum()) if "cnpj_limpo" in tratado else 0
        if cnpjs_validos <= 0:
            return False, "A base de clientes precisa conter CNPJ valido."
    elif chave == "produtos_mix":
        tratado = preparar_produtos_mix(bruto)
        eans_validos = int(tratado["ean_limpo"].dropna().astype(str).str.strip().ne("").sum()) if "ean_limpo" in tratado else 0
        produtos_validos = int(tratado["produto"].dropna().astype(str).str.strip().ne("").sum()) if "produto" in tratado else 0
        classificados = tratado[tratado["tipo_mix"].ne(TIPO_SEM_CLASSIFICACAO)] if "tipo_mix" in tratado else pd.DataFrame()
        if eans_validos <= 0:
            return False, "A coluna EAN nao foi encontrada ou nao possui valores validos."
        if produtos_validos <= 0:
            return False, "A coluna Produto nao foi encontrada ou esta vazia."
        if classificados.empty:
            return False, "Todos os produtos ficaram SEM CLASSIFICACAO."
    elif chave == "produtos_mercado_farma":
        coluna = "ean" if "ean" in base.columns else base.columns[0] if len(base.columns) else ""
        eans = base[coluna].dropna().astype(str).map(normalizar_ean) if coluna else pd.Series(dtype=str)
        if int(eans[eans.ne("")].nunique()) <= 0:
            return False, "A planilha precisa conter EANs validos."
    elif chave == "mercado_farma":
        if "ean" not in base.columns:
            return False, "A base Mercado Farma precisa conter EAN."
        if not any(coluna in base.columns for coluna in ["produto", "nome_do_produto"]):
            return False, "A base Mercado Farma precisa conter Produto."
        if "distribuidora" not in base.columns:
            return False, "A base Mercado Farma precisa conter Distribuidora."
    return True, ""


def tratar_bases(raw: BasesRaw) -> dict[str, pd.DataFrame | list[str]]:
    avisos: list[str] = []
    avisos.extend(validar_colunas_esperadas(raw.bussola, COLUNAS_BUSSOLA, "bussola.xlsx"))
    avisos.extend(validar_colunas_esperadas(raw.painel, COLUNAS_PAINEL, "PAINEL EQUIPE NORTE.xlsx"))
    if raw.acoes.empty:
        avisos.append("template_acoes_promocionais.xlsx: sem acoes cadastradas.")
    if raw.produtos_mix.empty:
        avisos.append("template_produtos_mix.xlsx: sem produtos classificados. Produtos vendidos ficarao como SEM CLASSIFICACAO.")
    else:
        avisos.extend(validar_colunas_esperadas(raw.produtos_mix, COLUNAS_PRODUTOS_MIX, "template_produtos_mix.xlsx"))

    clientes = preparar_painel_equipe(raw.painel)
    produtos_mix = preparar_produtos_mix(raw.produtos_mix)
    acoes = preparar_acoes(raw.acoes)
    vendas = preparar_base_vendas(raw.bussola, clientes, produtos_mix)

    return {
        "vendas": vendas,
        "clientes": clientes,
        "produtos_mix": produtos_mix,
        "acoes": acoes,
        "avisos": avisos,
        "raw_bussola": raw.bussola,
        "raw_bussola_historico": raw.bussola_historico,
        "raw_painel": raw.painel,
        "raw_acoes": raw.acoes,
        "raw_produtos_mix": raw.produtos_mix,
        "mercado_farma": raw.mercado_farma,
        "produtos_mercado_farma": raw.produtos_mercado_farma,
    }


def carregar_dados_tratados_de_arquivos(root_dir: Path | str) -> dict[str, pd.DataFrame | list[str]]:
    return tratar_bases(carregar_bases_de_arquivos(root_dir))


def compactar_upload_painel(conteudo: bytes) -> tuple[bytes, int]:
    painel_raw = ler_painel_bytes(conteudo)
    painel = preparar_painel_equipe(painel_raw)
    colunas_compactas_padrao = COLUNAS_PAINEL + COLUNAS_CONTATO + ["cnpj_limpo", "grupo_sip", "cliente_ativo"]
    if painel.empty:
        compacto = pd.DataFrame(columns=colunas_compactas_padrao)
    else:
        painel = painel[painel["cnpj_limpo"].apply(lambda cnpj: bool(normalizar_cnpj(cnpj) and normalizar_cnpj(cnpj) != "0" * 14))]
        painel = painel.drop_duplicates("cnpj_limpo", keep="first").reset_index(drop=True)
        colunas_compactas = [col for col in colunas_compactas_padrao if col in painel.columns]
        compacto = painel[colunas_compactas].copy()

    saida = BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        compacto.to_excel(writer, sheet_name="Planilha1", index=False)
    return saida.getvalue(), int(len(compacto.index))


def deduplicar_bussola_bytes(conteudo: bytes) -> bytes:
    bruto = ler_excel_bytes(conteudo, ABAS_PADRAO["bussola"])
    deduplicado = deduplicar_exportacao_bussola(bruto)
    saida = BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        deduplicado.to_excel(writer, sheet_name=ABAS_PADRAO["bussola"], index=False)
    return saida.getvalue()

