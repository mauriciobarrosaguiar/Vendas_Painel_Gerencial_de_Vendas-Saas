from __future__ import annotations

from html import escape
from io import BytesIO
import os
from uuid import uuid4

import pandas as pd
import requests
import streamlit as st

from src import github_actions as gha
from src import mercado_farma as mf
from src.configuracoes import carregar_login_bussola
from src.layout import card_metrica, dataframe_com_download, titulo_pagina
from src.loader import carregar_dados_tratados, registrar_upload
from src.status_bases import formatar_ultima_atualizacao
from src.tratamento import formatar_moeda, normalizar_cnpj, normalizar_ean


def desconto_texto(valor: object) -> str:
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0.0
    return f"{numero * 100:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def _texto(valor: object, padrao: str = "-") -> str:
    texto = "" if valor is None or pd.isna(valor) else str(valor).strip()
    return texto or padrao


def _html(valor: object, padrao: str = "-") -> str:
    return escape(_texto(valor, padrao))


def tabela_mercado_sem_consultor(df: pd.DataFrame) -> pd.DataFrame:
    tabela = mf.formatar_tabela_mercado(df)
    return tabela.drop(columns=["Consultor"], errors="ignore")


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _formatar_aba_excel_mercado_app(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0B5D3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    larguras: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            valor = "" if cell.value is None else str(cell.value)
            larguras[cell.column_letter] = min(max(larguras.get(cell.column_letter, 0), len(valor) + 2), 45)
            cell.border = border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, _ in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = max(larguras.get(letra, 12), 10)
        cabecalho = ws.cell(row=1, column=idx).value
        if cabecalho == "Desconto":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "0.00%"
        elif cabecalho in {"PF Dist.", "PF Fábrica", "Preço com imposto", "Preço sem imposto"}:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0.00"
        elif cabecalho == "Estoque":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0"
        elif cabecalho == "Atualizado em":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "dd/mm/yyyy hh:mm"


def _tabela_excel_mercado_app(df: pd.DataFrame) -> pd.DataFrame:
    base = mf.preparar_mercado_farma(df)
    if not base.empty:
        status_erro = base["status"].astype(str).str.strip().str.upper().isin({"ERRO", "NAO ENCONTRADO"})
        sem_produto = base["produto"].astype(str).str.strip().eq("")
        sem_distribuidora = base["distribuidora"].astype(str).str.strip().eq("")
        sem_valor = (base["estoque"].fillna(0).astype(float) <= 0) & (base["preco_sem_imposto"].fillna(0).astype(float) <= 0)
        erro_generico = base["erro"].astype(str).str.strip().str.lower().isin({"", "message:", "message"})
        mask_nao_encontrado = status_erro & sem_produto & sem_distribuidora & sem_valor & erro_generico
        base.loc[mask_nao_encontrado, "produto"] = "Produto nao encontrado"
        base.loc[mask_nao_encontrado, "status"] = "NAO ENCONTRADO"
        base.loc[mask_nao_encontrado, "erro"] = "EAN nao encontrado no Mercado Farma"

    colunas = {
        "uf": "UF",
        "cnpj_referencia": "CNPJ referência",
        "ean": "EAN",
        "produto": "Produto",
        "distribuidora": "Distribuidora",
        "estoque": "Estoque",
        "desconto": "Desconto",
        "pf_dist": "PF Dist.",
        "pf_fabrica": "PF Fábrica",
        "preco_com_imposto": "Preço com imposto",
        "preco_sem_imposto": "Preço sem imposto",
        "data_atualizacao": "Atualizado em",
        "status": "Status",
        "erro": "Erro",
    }
    tabela = base.rename(columns=colunas)
    return tabela[list(colunas.values())]


def _ordenar_tabela_excel_mercado_app(base: pd.DataFrame) -> pd.DataFrame:
    if base.empty:
        return base
    tabela = base.copy()
    status = tabela["Status"].fillna("").astype(str).str.strip().str.upper() if "Status" in tabela.columns else pd.Series("", index=tabela.index)
    tabela["_ordem_status"] = 2
    tabela.loc[status.eq("OK"), "_ordem_status"] = 0
    tabela.loc[status.eq("NAO ENCONTRADO"), "_ordem_status"] = 1
    ordenacao = [col for col in ["UF", "_ordem_status", "Produto", "EAN", "Distribuidora"] if col in tabela.columns]
    return tabela.sort_values(ordenacao, kind="stable").drop(columns=["_ordem_status"], errors="ignore").reset_index(drop=True)


def excel_mercado_farma_por_uf_app(df: pd.DataFrame) -> bytes:
    if hasattr(mf, "excel_mercado_farma_por_uf"):
        return mf.excel_mercado_farma_por_uf(df)

    base = _ordenar_tabela_excel_mercado_app(_tabela_excel_mercado_app(df))

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if base.empty or "UF" not in base.columns:
            base.to_excel(writer, sheet_name="Mercado Farma", index=False)
            _formatar_aba_excel_mercado_app(writer.book["Mercado Farma"])
        else:
            ufs = sorted(uf for uf in base["UF"].dropna().astype(str).str.strip().str.upper().unique().tolist() if uf)
            if not ufs:
                base.to_excel(writer, sheet_name="SEM_UF", index=False)
                _formatar_aba_excel_mercado_app(writer.book["SEM_UF"])
            for uf in ufs:
                df_uf = base[base["UF"].astype(str).str.upper().eq(uf)].copy()
                df_uf.to_excel(writer, sheet_name=uf[:31], index=False)
                _formatar_aba_excel_mercado_app(writer.book[uf[:31]])
    return buffer.getvalue()


def botao_download_mercado_excel(df: pd.DataFrame, nome_arquivo: str, rotulo: str, key: str) -> None:
    base = mf.preparar_mercado_farma(df)
    st.download_button(
        rotulo,
        data=excel_mercado_farma_por_uf_app(base),
        file_name=nome_arquivo,
        mime=MIME_XLSX,
        width="stretch",
        disabled=base.empty,
        key=key,
    )


def renderizar_downloads_mercado(base: pd.DataFrame) -> None:
    dados = mf.preparar_mercado_farma(base)
    with st.expander("Downloads Mercado Farma", expanded=False):
        st.caption("Baixe todas as UFs em abas separadas ou gere uma planilha individual por UF.")
        botao_download_mercado_excel(
            dados,
            "mercado_farma_todas_ufs.xlsx",
            "Baixar todas as UFs em abas separadas",
            "mf_download_todas_ufs",
        )
        if dados.empty:
            return

        ufs_download = sorted(dados["uf"].dropna().astype(str).str.strip().str.upper().replace("", pd.NA).dropna().unique().tolist())
        if not ufs_download:
            return

        st.markdown("<span class='pill-note'>Planilha por UF</span>", unsafe_allow_html=True)
        colunas = st.columns(min(3, len(ufs_download)))
        for idx, uf in enumerate(ufs_download):
            df_uf = dados[dados["uf"].astype(str).str.upper().eq(uf)].copy()
            with colunas[idx % len(colunas)]:
                botao_download_mercado_excel(df_uf, f"mercado_farma_{uf}.xlsx", f"Baixar {uf}", f"mf_download_uf_{uf}")


def secret_app(nome: str, padrao: str = "") -> str:
    try:
        if nome in st.secrets:
            return str(st.secrets[nome])
    except Exception:
        pass
    return str(os.environ.get(nome, padrao) or padrao)


def mascarar_usuario_app(usuario: object) -> str:
    if hasattr(mf, "mascarar_usuario"):
        return str(mf.mascarar_usuario(usuario))
    texto = _texto(usuario, "")
    if not texto:
        return ""
    if "@" in texto:
        nome, dominio = texto.split("@", 1)
        return f"{nome[:2]}***{nome[-1:]}@{dominio}" if len(nome) > 2 else f"{nome[:1]}***@{dominio}"
    return f"{texto[:2]}***{texto[-2:]}" if len(texto) > 4 else f"{texto[:1]}***"


def carregar_credenciais_mercadofarma_app(login: dict | None) -> dict[str, object]:
    if hasattr(mf, "carregar_credenciais_mercadofarma"):
        return mf.carregar_credenciais_mercadofarma(login)

    login = login if isinstance(login, dict) else {}
    usuario = secret_app("MERCADOFARMA_USUARIO")
    senha = secret_app("MERCADOFARMA_SENHA")
    for dados in [login.get("mercadofarma", {}), login.get("mercado_farma", {}), login.get("gd", {})]:
        if usuario and senha:
            break
        if isinstance(dados, dict):
            usuario = usuario or _texto(dados.get("usuario"), "")
            senha = senha or _texto(dados.get("senha"), "")

    faltantes = []
    if not usuario:
        faltantes.append("MERCADOFARMA_USUARIO")
    if not senha:
        faltantes.append("MERCADOFARMA_SENHA")
    return {
        "usuario": usuario,
        "senha": senha,
        "usuario_mascarado": mascarar_usuario_app(usuario),
        "configurado": not faltantes,
        "faltantes": faltantes,
    }


def alvos_mercadofarma_por_uf_app(clientes: pd.DataFrame, usuario_gd: str, senha_gd: str) -> list[dict[str, str]]:
    if hasattr(mf, "alvos_mercadofarma_por_uf"):
        return mf.alvos_mercadofarma_por_uf(clientes, usuario_gd, senha_gd)

    if clientes is None or clientes.empty:
        return []
    base = clientes.copy()
    for coluna in ["uf", "cnpj_limpo"]:
        if coluna not in base.columns:
            base[coluna] = ""
    if "cliente_ativo" in base.columns:
        base = base[base["cliente_ativo"].fillna(True)].copy()
    base["uf"] = base["uf"].astype(str).str.strip().str.upper()
    base["cnpj_limpo"] = base["cnpj_limpo"].apply(normalizar_cnpj)
    valid_ufs = getattr(mf, "VALID_UFS", {"MA", "MT", "PA", "PI", "TO"})
    base = base[base["uf"].isin(valid_ufs) & base["cnpj_limpo"].str.len().eq(14)].copy()
    alvos = []
    for uf, grupo_uf in base.sort_values(["uf", "cnpj_limpo"]).groupby("uf", dropna=False):
        cnpj = str(grupo_uf["cnpj_limpo"].iloc[0])
        alvos.append({"consultor": "GD", "uf": str(uf), "cnpj": cnpj, "usuario": usuario_gd, "senha": senha_gd})
    return sorted(alvos, key=lambda item: item["uf"])


def disparar_mercado_farma_app(ufs: list[str], limite_eans: int, usuario_gd: str, senha_gd: str) -> None:
    persistence_key = secret_app("PERSISTENCE_KEY")
    try:
        gha.disparar_mercado_farma(
            ufs,
            limite_eans,
            mercadofarma_usuario=usuario_gd,
            mercadofarma_senha=senha_gd,
            persistence_key=persistence_key,
        )
        return
    except TypeError as exc:
        if "mercadofarma_usuario" not in str(exc) and "persistence_key" not in str(exc):
            raise

    token = secret_app("GITHUB_TOKEN")
    repo = secret_app("GITHUB_REPO", "mauriciobarrosaguiar/painel-comercial-equipe-norte")
    branch = secret_app("GITHUB_BRANCH", "main")
    if not token or not repo:
        raise RuntimeError("Configure GITHUB_TOKEN e GITHUB_REPO nos Secrets para disparar a atualização.")

    ufs_txt = ",".join(str(uf).strip().upper() for uf in ufs if str(uf).strip())
    payload = {
        "ref": branch,
        "inputs": {
            "acao": "atualizar_mercadofarma_paralelo",
            "ufs": ufs_txt,
            "uf": ufs_txt,
            "limite_eans": str(int(limite_eans or 0)),
            "headless": "true",
            "mercadofarma_usuario": str(usuario_gd or ""),
            "mercadofarma_senha": str(senha_gd or ""),
            "persistence_key": persistence_key,
            "command_id": uuid4().hex,
        },
    }
    headers = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    url = f"https://api.github.com/repos/{repo}/actions/workflows/mercadofarma.yml/dispatches"
    resp = requests.post(url, headers=headers, json=payload, timeout=30)
    if resp.status_code not in {204, 201}:
        raise RuntimeError(f"A atualização não foi aceita ({resp.status_code}): {resp.text[:500]}")


def produto_card_distribuidora(grupo: pd.DataFrame, key: str) -> None:
    opcoes = grupo.sort_values(["preco_sem_imposto", "estoque"], ascending=[True, False]).reset_index(drop=True)
    if opcoes.empty:
        return

    with st.container(border=True):
        primeiro = opcoes.iloc[0]
        st.markdown(
            f"""
            <div class="produto-top">
                <span class="desconto-badge">{desconto_texto(primeiro.get('desconto', 0))}</span>
                <span class="produto-meta">{_html(primeiro.get('uf'))}</span>
            </div>
            <div class="produto-nome">{_html(primeiro.get('produto'), 'Produto sem descrição')}</div>
            <div class="produto-meta">EMS Genéricos &nbsp; | &nbsp; {_html(primeiro.get('ean'))}</div>
            """,
            unsafe_allow_html=True,
        )
        st.caption("Distribuidora")
        if len(opcoes) > 1:
            def rotulo(indice: int) -> str:
                item = opcoes.iloc[indice]
                dist = _texto(item.get("distribuidora"), "Distribuidora não identificada")
                preco = formatar_moeda(item.get("preco_sem_imposto", 0))
                estoque = int(float(item.get("estoque", 0) or 0))
                return f"{dist} | {preco} | {estoque} un."

            escolha = st.selectbox(
                "Distribuidora do produto",
                list(range(len(opcoes))),
                format_func=rotulo,
                key=key,
                label_visibility="collapsed",
            )
        else:
            escolha = 0
            st.markdown(
                f"<span class='pill-note'>{_html(opcoes.iloc[0].get('distribuidora'), 'Distribuidora não identificada')}</span>",
                unsafe_allow_html=True,
            )

        item = opcoes.iloc[int(escolha)]
        preco = float(item.get("preco_sem_imposto", 0) or 0)
        preco_com = float(item.get("preco_com_imposto", 0) or 0)
        pf_dist = float(item.get("pf_dist", 0) or 0)
        estoque = int(float(item.get("estoque", 0) or 0))
        st.markdown(
            f"""
            <div class="preco-box">
                <div>
                    <div class="preco-dist">{_html(item.get('distribuidora'), 'Distribuidora não identificada')}</div>
                    <div class="preco-estoque">{estoque} un. disponíveis</div>
                </div>
                <div>
                    <div class="preco-secundario">PF Dist.: {formatar_moeda(pf_dist)}</div>
                    <div class="preco-principal">{formatar_moeda(preco)}</div>
                    <div class="preco-secundario">Com imposto: {formatar_moeda(preco_com)}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def painel_status_extracao(estado: dict) -> None:
    status = str(estado.get("status") or "parado").upper()
    total = int(estado.get("total_passos", 0) or 0)
    processados = int(estado.get("processados", 0) or 0)
    percentual = 0 if total <= 0 else min(max(processados / total, 0), 1)
    st.progress(percentual)
    st.markdown(
        " ".join(
            [
                f"<span class='pill-note'>Status: {escape(status)}</span>",
                f"<span class='pill-note'>Processados: {processados}/{total}</span>",
                f"<span class='pill-note'>UF atual: {escape(str(estado.get('current_uf') or '-'))}</span>",
                f"<span class='pill-note'>EAN atual: {escape(str(estado.get('current_ean') or '-'))}</span>",
            ]
        ),
        unsafe_allow_html=True,
    )
    if estado.get("mensagem"):
        st.caption(str(estado["mensagem"]))
    if estado.get("erro"):
        st.error(str(estado["erro"]))
    logs = estado.get("logs", [])
    if logs:
        linhas = []
        for item in logs[-18:]:
            texto = str(item)
            if " / " in texto and ": " in texto:
                inicio, resto = texto.split(" - ", 1) if " - " in texto else ("", texto)
                partes = resto.split(" / ", 1)
                if len(partes) == 2:
                    texto = f"{inicio} - UF {partes[1]}" if inicio else f"UF {partes[1]}"
            linhas.append(texto)
        st.code("\n".join(linhas), language="text")


def tabela_status_consolidado(status: dict) -> pd.DataFrame:
    itens = status.get("status", []) if isinstance(status, dict) else []
    if not isinstance(itens, list):
        return pd.DataFrame()
    linhas = []
    for item in itens:
        if not isinstance(item, dict):
            continue
        linhas.append(
            {
                "UF": item.get("uf", ""),
                "Status": item.get("status", ""),
                "CNPJ referência": item.get("cnpj_referencia", ""),
                "Produtos": item.get("total_produtos", 0),
                "Erro": item.get("erro", ""),
            }
        )
    return pd.DataFrame(linhas)


def renderizar_execucoes_actions(runs: list[dict]) -> None:
    if not runs:
        return
    st.markdown("<span class='pill-note'>Últimas execuções</span>", unsafe_allow_html=True)
    for run in runs:
        falhou = run.get("conclusao") == "failure" or bool(run.get("ufs_com_erro"))
        with st.container(border=True):
            c1, c2, c3, c4 = st.columns(4)
            c1.caption("Criada em")
            c1.markdown(f"**{run.get('criada_em', '-')}**")
            c2.caption("Status")
            c2.markdown(f"**{run.get('status_pt', '-')}**")
            c3.caption("Resultado")
            c3.markdown(f"**{run.get('conclusao_pt', '-')}**")
            c4.caption("UF")
            c4.markdown(f"**{run.get('uf', '-')}**")

            d1, d2 = st.columns(2)
            d1.caption(f"Branch: {run.get('branch', '-')}")
            d2.caption(f"Ação executada: {run.get('acao', '-')}")

            links = []
            if run.get("url"):
                links.append(f"[Abrir execução]({run['url']})")
            if run.get("jobs_url"):
                links.append(f"[API dos jobs]({run['jobs_url']})")
            if links:
                st.markdown(" | ".join(links))

            if falhou:
                chave = f"mf_ver_erro_{run.get('id')}"
                if st.button("Ver detalhes do erro" if not st.session_state.get(chave) else "Ocultar detalhes do erro", key=f"btn_{chave}", width="stretch"):
                    st.session_state[chave] = not bool(st.session_state.get(chave))
                if st.session_state.get(chave):
                    jobs_falhos = [job for job in run.get("jobs", []) if job.get("conclusao") == "failure"]
                    if jobs_falhos:
                        for job in jobs_falhos:
                            st.markdown(
                                f"**{job.get('nome', 'Job')}** | Status: **{job.get('status_pt', '-')}** | Resultado: **{job.get('conclusao_pt', '-')}**"
                            )
                            if job.get("html_url"):
                                st.markdown(f"[Abrir logs da UF {job.get('uf', '-')}]({job['html_url']})")
                            if job.get("logs_url"):
                                st.caption(f"logs_url: {job['logs_url']}")
                            if job.get("erro_resumo"):
                                st.code(job["erro_resumo"], language="text")
                    elif run.get("erro_resumo"):
                        st.code(run["erro_resumo"], language="text")
                    else:
                        st.info("A execução falhou, mas o GitHub ainda não liberou detalhes de log pela API.")


def configurar_desconto_adicional(mercado_base: pd.DataFrame) -> dict:
    config = mf.carregar_descontos_adicionais()
    with st.expander("Desconto adicional por distribuidora", expanded=False):
        if mercado_base.empty:
            st.info("Extraia ou importe o Mercado Farma para cadastrar desconto adicional.")
            return config
        distribuidoras = sorted(mercado_base["distribuidora"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist())
        if not distribuidoras:
            st.info("Nenhuma distribuidora encontrada na base atual.")
            return config
        dist = st.selectbox("Distribuidora", distribuidoras, key="mf_desconto_dist")
        regras = config.setdefault("distribuidoras", {})
        regra = regras.get(dist, {})
        percentual_atual = float(regra.get("percentual", 0) or 0)
        percentual_visual = percentual_atual * 100 if percentual_atual <= 1 else percentual_atual
        percentual = st.number_input("Desconto adicional (%)", min_value=0.0, max_value=100.0, step=0.5, value=float(percentual_visual), key="mf_desconto_pct")

        produtos_dist = mercado_base[mercado_base["distribuidora"].astype(str).eq(dist)].copy()
        produtos_dist = produtos_dist[["ean", "produto"]].drop_duplicates("ean").sort_values("produto")
        mapa_label_ean = {
            f"{_texto(row.produto, 'Produto sem descrição')} | {row.ean}": str(row.ean)
            for row in produtos_dist.itertuples(index=False)
        }
        eans_sem = set(str(ean) for ean in regra.get("eans_sem_desconto", []))
        default_labels = [label for label, ean in mapa_label_ean.items() if ean in eans_sem]
        selecionados = st.multiselect(
            "Produtos sem desconto adicional nesta distribuidora",
            list(mapa_label_ean.keys()),
            default=default_labels,
            key="mf_desconto_excecoes",
        )
        c1, c2 = st.columns(2)
        if c1.button("Salvar desconto adicional", width="stretch"):
            regras[dist] = {
                "percentual": float(percentual) / 100,
                "eans_sem_desconto": [normalizar_ean(mapa_label_ean[label]) for label in selecionados],
            }
            mf.salvar_descontos_adicionais(config)
            st.success("Desconto adicional salvo.")
            st.rerun()
        if c2.button("Remover desconto da distribuidora", width="stretch", disabled=dist not in regras):
            regras.pop(dist, None)
            mf.salvar_descontos_adicionais(config)
            st.success("Desconto adicional removido.")
            st.rerun()

        if regras:
            resumo = [
                {
                    "Distribuidora": nome,
                    "Desconto adicional": f"{float(regra.get('percentual', 0) or 0) * 100:.2f}%",
                    "Produtos sem adicional": len(regra.get("eans_sem_desconto", [])),
                }
                for nome, regra in regras.items()
                if isinstance(regra, dict)
            ]
            st.dataframe(pd.DataFrame(resumo), width="stretch", hide_index=True)
    return config


dados = carregar_dados_tratados()
clientes = dados["clientes"]
produtos_mercado = dados["produtos_mercado_farma"]

titulo_pagina("Mercado Farma / UF", "Preços e estoque por UF da carteira")

mercado_original = mf.mercado_farma_atual()
descontos_config = mf.carregar_descontos_adicionais()
mercado = mf.aplicar_descontos_adicionais(mercado_original, descontos_config)
login = carregar_login_bussola()
credencial_gd = carregar_credenciais_mercadofarma_app(login)
usuario_gd = str(credencial_gd.get("usuario", ""))
senha_gd = str(credencial_gd.get("senha", ""))
credenciais = [{"consultor": "GD", "usuario": usuario_gd, "senha": senha_gd}] if credencial_gd.get("configurado") else []
alvos = alvos_mercadofarma_por_uf_app(clientes, usuario_gd, senha_gd)
ufs_carteira = set(mf.ufs_validas_clientes(clientes))
ufs_alvos = sorted({alvo["uf"] for alvo in alvos} or ufs_carteira)
ufs_sem_cnpj = sorted(ufs_carteira - {alvo["uf"] for alvo in alvos})

st.markdown(f"<span class='pill-note'>Última atualização consolidada: {formatar_ultima_atualizacao('mercado_farma')}</span>", unsafe_allow_html=True)

with st.expander("Extração Mercado Farma", expanded=False):
    st.caption("Cada UF usa um CNPJ de referência da própria carteira. A extração usa somente o acesso GD do Mercado Farma.")
    if credencial_gd.get("configurado"):
        st.markdown(
            f"<span class='pill-note'>Acesso GD configurado: {escape(str(credencial_gd.get('usuario_mascarado') or 'usuario oculto'))}</span>",
            unsafe_allow_html=True,
        )
    else:
        faltantes = ", ".join(str(item) for item in credencial_gd.get("faltantes", []))
        st.warning(f"Configure o acesso GD do Mercado Farma nos Secrets do Streamlit/GitHub: {faltantes}.")
    if alvos:
        tabela_alvos = pd.DataFrame([{"UF": item["uf"], "CNPJ referência": item["cnpj"]} for item in alvos])
        st.dataframe(tabela_alvos, width="stretch", height=170, hide_index=True)
    else:
        st.info("Não encontrei CNPJ referência ativo na carteira para montar a extração por UF.")
    if ufs_sem_cnpj:
        st.warning("UFs na carteira sem CNPJ referência ativo: " + ", ".join(ufs_sem_cnpj))

    eans = mf.obter_eans_para_consulta(produtos_mercado)
    st.markdown(
        f"<span class='pill-note'>Lista produtos.xlsx: {len(eans)} EANs</span>"
        f"<span class='pill-note'>Atualização da lista: {formatar_ultima_atualizacao('produtos_mercado_farma')}</span>",
        unsafe_allow_html=True,
    )
    upload_eans = st.file_uploader("Atualizar planilha produtos.xlsx com EANs", type=["xlsx"], key="upload_produtos_mercado_farma")
    if upload_eans is not None:
        registrar_upload("produtos_mercado_farma", upload_eans)
        st.cache_data.clear()
        st.success("Lista produtos.xlsx salva para as próximas extrações.")
        st.rerun()

    ufs_para_rodar = st.multiselect("UFs para atualizar", ufs_alvos, default=ufs_alvos, key="mf_ufs_rodar")
    limite_eans = st.number_input("Limite de EANs para teste (0 = todos)", min_value=0, step=10, value=0)

    col_git1, col_git2 = st.columns(2)
    if col_git1.button("Atualizar UFs Selecionadas", width="stretch", disabled=not bool(ufs_para_rodar)):
        try:
            disparar_mercado_farma_app(
                ufs_para_rodar,
                int(limite_eans or 0),
                usuario_gd,
                senha_gd,
            )
            st.success("Atualização iniciada. Acompanhe o status abaixo.")
        except Exception as exc:
            st.error(f"Não consegui iniciar a atualização: {exc}")
    if col_git2.button("Atualizar Todas as UFs", width="stretch", disabled=not bool(ufs_alvos)):
        try:
            disparar_mercado_farma_app(
                ufs_alvos,
                int(limite_eans or 0),
                usuario_gd,
                senha_gd,
            )
            st.success("Atualização iniciada para todas as UFs.")
        except Exception as exc:
            st.error(f"Não consegui iniciar a atualização: {exc}")

    status_consolidado = mf.carregar_status_consolidado()
    status_tabela = tabela_status_consolidado(status_consolidado)
    if not status_tabela.empty:
        st.markdown("<span class='pill-note'>Status do consolidado</span>", unsafe_allow_html=True)
        st.dataframe(status_tabela, width="stretch", hide_index=True)

    runs = gha.listar_execucoes_mercado_farma(5)
    if gha.limpar_cache_se_mercado_farma_finalizou(runs):
        st.info("Execução do GitHub Actions finalizada. Cache limpo para recarregar o consolidado Mercado Farma.")
        st.rerun()
    if runs:
        renderizar_execucoes_actions(runs)

    with st.expander("Extração local de apoio", expanded=False):
        estado = mf.carregar_estado_extracao()
        painel_status_extracao(estado)
        headless = st.toggle("Rodar navegador oculto", value=True, key="mercado_headless")
        rodando = estado.get("status") == "rodando" and estado.get("thread_alive")
        pode_retomar = estado.get("status") in {"erro", "cancelado", "interrompido"}
        col1, col2, col3 = st.columns(3)
        local_desabilitado = rodando or not bool(ufs_para_rodar) or not bool(credenciais)
        if col1.button("Iniciar extração local", width="stretch", disabled=local_desabilitado):
            try:
                mf.iniciar_extracao_background(
                    credenciais,
                    clientes,
                    produtos_mercado,
                    headless=headless,
                    limite_eans=int(limite_eans) if limite_eans else None,
                    retomar=False,
                    ufs=ufs_para_rodar,
                )
                st.success("Extração local iniciada.")
                st.rerun()
            except Exception as exc:
                st.error(f"Falha ao iniciar extração local: {exc}")

        if col2.button("Retomar local", width="stretch", disabled=rodando or not pode_retomar or not bool(credenciais)):
            try:
                mf.iniciar_extracao_background(
                    credenciais,
                    clientes,
                    produtos_mercado,
                    headless=headless,
                    limite_eans=int(limite_eans) if limite_eans else None,
                    retomar=True,
                    ufs=ufs_para_rodar,
                )
                st.success("Extração local retomada.")
                st.rerun()
            except Exception as exc:
                st.error(f"Falha ao retomar extração local: {exc}")

        if col3.button("Cancelar local", width="stretch", disabled=not rodando):
            mf.cancelar_extracao_background()
            st.warning("Cancelamento solicitado.")
            st.rerun()

    upload = st.file_uploader("Importar planilha Mercado Farma", type=["xlsx"], key="upload_mercado_farma")
    if upload is not None:
        registrar_upload("mercado_farma", upload)
        st.cache_data.clear()
        st.success("Planilha Mercado Farma salva.")
        st.rerun()

if mercado.empty:
    st.info("Ainda não existe base do Mercado Farma salva. Extraia pelo botão acima ou importe uma planilha.")
    st.stop()

preco_valido = pd.to_numeric(mercado["preco_sem_imposto"], errors="coerce").fillna(0) > 0
estoque_valido = pd.to_numeric(mercado["estoque"], errors="coerce").fillna(0) > 0
mercado_valido = mercado[preco_valido & estoque_valido].copy()

configurar_desconto_adicional(mf.preparar_mercado_farma(mercado_original))

mf_metricas = mercado_valido.copy()
m1, m2, m3, m4 = st.columns(4)
with m1:
    card_metrica("Produtos com preço", str(int(mf_metricas["ean"].nunique())))
with m2:
    card_metrica("UFs", str(int(mf_metricas["uf"].nunique())))
with m3:
    card_metrica("Distribuidoras", str(int(mf_metricas["distribuidora"].nunique())))
with m4:
    estoque_total = int(pd.to_numeric(mf_metricas["estoque"], errors="coerce").fillna(0).sum())
    card_metrica("Estoque total", f"{estoque_total:,}".replace(",", "."))

renderizar_downloads_mercado(mercado)

total_melhores = len(mf.melhor_preco_por_ean(mercado_valido))
with st.expander(f"Melhores preços — {total_melhores} produtos encontrados", expanded=False):
    f1, f2, f3, f4 = st.columns([1.6, 0.8, 1.2, 0.7])
    busca = f1.text_input("Buscar produto, EAN ou distribuidora", key="mf_busca_melhores")
    uf_sel = f2.multiselect("UF", sorted(mercado_valido["uf"].dropna().astype(str).unique().tolist()), key="mf_uf_melhores")
    distribuidora_sel = f3.multiselect(
        "Distribuidora",
        sorted(mercado_valido["distribuidora"].dropna().astype(str).unique().tolist()),
        key="mf_dist_melhores",
    )
    buscar = f4.button("Buscar", width="stretch", key="mf_botao_buscar")
    if buscar:
        st.session_state["mf_mostrar_melhores"] = True

    filtrado = mercado_valido.copy()
    if uf_sel:
        filtrado = filtrado[filtrado["uf"].isin(uf_sel)].copy()
    if distribuidora_sel:
        filtrado = filtrado[filtrado["distribuidora"].isin(distribuidora_sel)].copy()
    if busca:
        termo = busca.strip().lower()
        mask = (
            filtrado["produto"].astype(str).str.lower().str.contains(termo, na=False, regex=False)
            | filtrado["ean"].astype(str).str.lower().str.contains(termo, na=False, regex=False)
            | filtrado["distribuidora"].astype(str).str.lower().str.contains(termo, na=False, regex=False)
        )
        filtrado = filtrado[mask].copy()

    melhores = mf.melhor_preco_por_ean(filtrado)
    deve_mostrar = bool(st.session_state.get("mf_mostrar_melhores")) or bool(busca or uf_sel or distribuidora_sel)
    if not deve_mostrar:
        st.info("Use a busca ou os filtros e clique em Buscar para carregar os cards de melhores preços.")
    elif melhores.empty:
        st.info("Sem produtos com preço e estoque para os filtros selecionados.")
    else:
        limite_cards = min(len(melhores), 60)
        for fatia in [melhores.iloc[i : i + 3] for i in range(0, limite_cards, 3)]:
            cols = st.columns(3)
            for col, (_, item) in zip(cols, fatia.iterrows()):
                with col:
                    grupo = filtrado[(filtrado["uf"] == item["uf"]) & (filtrado["ean"] == item["ean"])].copy()
                    produto_card_distribuidora(grupo, f"dist_{item['uf']}_{item['ean']}_{int(item.name)}")

    if deve_mostrar:
        c1, c2 = st.columns(2)
        with c1:
            botao_download_mercado_excel(filtrado, "mercado_farma_por_uf.xlsx", "Extrair lista completa em Excel", "mf_download_lista_filtrada")
        with c2:
            botao_download_mercado_excel(melhores, "mercado_farma_melhores_precos.xlsx", "Extrair melhores preços em Excel", "mf_download_melhores")

        with st.expander("Tabela completa", expanded=False):
            dataframe_com_download(tabela_mercado_sem_consultor(filtrado), "mercado_farma_completo", altura=420)
