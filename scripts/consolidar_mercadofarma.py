from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
BASE_DIR = ROOT / "data" / "mercadofarma"
PARCIAIS_DIR = BASE_DIR / "parciais"
FINAL_PATH = BASE_DIR / "mercadofarma_consolidado.csv"
FINAL_XLSX_PATH = BASE_DIR / "mercadofarma_consolidado.xlsx"
STATUS_PATH = BASE_DIR / "status_mercadofarma.json"
TZ_BRASILIA = ZoneInfo("America/Sao_Paulo")
COLUNAS_MINIMAS = {"UF", "EAN", "CNPJ_REFERENCIA"}


def log(msg: str) -> None:
    print(msg, flush=True)


def agora_brasilia_iso() -> str:
    return datetime.now(TZ_BRASILIA).isoformat()


def ler_csv(path: Path) -> pd.DataFrame:
    candidatos: list[pd.DataFrame] = []
    for sep in [",", ";"]:
        try:
            df = pd.read_csv(path, dtype=str, sep=sep)
        except Exception:
            continue
        candidatos.append(df)
        if COLUNAS_MINIMAS.issubset(set(df.columns)):
            return df
    if candidatos:
        return max(candidatos, key=lambda item: len(item.columns))
    raise RuntimeError(f"Nao consegui ler {path}")


def arquivos_parciais() -> list[Path]:
    if not PARCIAIS_DIR.exists():
        return []
    return sorted(PARCIAIS_DIR.rglob("mercadofarma_*.csv"))


def arquivos_status() -> list[Path]:
    if not BASE_DIR.exists():
        return []
    return sorted(BASE_DIR.rglob("mercadofarma_*.json"))


def ler_status_ufs() -> dict[str, dict]:
    status_por_uf: dict[str, dict] = {}
    for path in arquivos_status():
        try:
            item = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            log(f"Aviso: status invalido em {path}: {exc}")
            continue
        if not isinstance(item, dict):
            continue
        uf = str(item.get("uf") or "").strip().upper()
        if not uf:
            continue
        anterior = status_por_uf.get(uf)
        atual_fim = str(item.get("finalizado_em") or item.get("iniciado_em") or "")
        anterior_fim = str(anterior.get("finalizado_em") or anterior.get("iniciado_em") or "") if anterior else ""
        if anterior is None or atual_fim >= anterior_fim:
            status_por_uf[uf] = item
    return status_por_uf


def normalizar_uf(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    if "UF" not in base.columns and "uf" in base.columns:
        base = base.rename(columns={"uf": "UF"})
    if "UF" not in base.columns:
        base["UF"] = ""
    base["UF"] = base["UF"].astype(str).str.strip().str.upper()
    return base


def uf_arquivo(path: Path, df: pd.DataFrame) -> str:
    if "UF" in df.columns and not df["UF"].dropna().empty:
        uf = str(df["UF"].dropna().astype(str).iloc[0]).strip().upper()
        if uf:
            return uf
    nome = path.stem.replace("mercadofarma_", "").strip().upper()
    return nome if len(nome) == 2 else ""


def csv_valido(df: pd.DataFrame) -> tuple[bool, str]:
    if df.empty:
        return False, "arquivo sem linhas"
    faltantes = sorted(COLUNAS_MINIMAS - set(df.columns))
    if faltantes:
        return False, "colunas ausentes: " + ", ".join(faltantes)
    if df["UF"].astype(str).str.strip().eq("").all():
        return False, "UF vazia"
    if df["EAN"].astype(str).str.strip().eq("").all():
        return False, "EAN vazio"
    return True, ""


def nome_aba_uf(valor: object) -> str:
    texto = "" if valor is None else str(valor).strip().upper()
    texto = "".join(char for char in texto if char not in r"[]:*?/\\")
    return (texto or "SEM_UF")[:31]


def formatar_aba_excel(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0B5D3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    larguras: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            valor = "" if cell.value is None else str(cell.value)
            larguras[cell.column_letter] = min(max(larguras.get(cell.column_letter, 0), len(valor) + 2), 45)
            cell.border = border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for idx in range(1, ws.max_column + 1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = max(larguras.get(letra, 12), 10)
        cabecalho = ws.cell(row=1, column=idx).value
        if cabecalho == "DESCONTO":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "0.00%"
        elif cabecalho in {"PF_DIST", "PF_FABRICA", "PRECO_COM_IMPOSTO", "PRECO_SEM_IMPOSTO"}:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0.00"
        elif cabecalho == "ESTOQUE":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0"


def salvar_excel_por_uf(df: pd.DataFrame) -> None:
    base = normalizar_uf(df)
    if base.empty:
        return
    for coluna in ["ESTOQUE", "DESCONTO", "PF_DIST", "PF_FABRICA", "PRECO_COM_IMPOSTO", "PRECO_SEM_IMPOSTO"]:
        if coluna in base.columns:
            base[coluna] = pd.to_numeric(base[coluna], errors="coerce").fillna(0)
    if {"STATUS", "PRODUTO", "DISTRIBUIDORA", "ESTOQUE", "PRECO_SEM_IMPOSTO", "ERRO"}.issubset(base.columns):
        status_erro = base["STATUS"].astype(str).str.strip().str.upper().isin({"ERRO", "NAO ENCONTRADO"})
        sem_produto = base["PRODUTO"].fillna("").astype(str).str.strip().eq("")
        sem_distribuidora = base["DISTRIBUIDORA"].fillna("").astype(str).str.strip().eq("")
        sem_valor = (base["ESTOQUE"] <= 0) & (base["PRECO_SEM_IMPOSTO"] <= 0)
        erro_generico = base["ERRO"].fillna("").astype(str).str.strip().str.lower().isin({"", "message:", "message"})
        mask_nao_encontrado = status_erro & sem_produto & sem_distribuidora & sem_valor & erro_generico
        base.loc[mask_nao_encontrado, "PRODUTO"] = "Produto nao encontrado"
        base.loc[mask_nao_encontrado, "STATUS"] = "NAO ENCONTRADO"
        base.loc[mask_nao_encontrado, "ERRO"] = "EAN nao encontrado no Mercado Farma"
    if "STATUS" in base.columns:
        status = base["STATUS"].fillna("").astype(str).str.strip().str.upper()
        base["_ORDEM_STATUS"] = 2
        base.loc[status.eq("OK"), "_ORDEM_STATUS"] = 0
        base.loc[status.eq("NAO ENCONTRADO"), "_ORDEM_STATUS"] = 1
    ordenacao = [col for col in ["UF", "_ORDEM_STATUS", "PRODUTO", "EAN", "DISTRIBUIDORA"] if col in base.columns]
    if ordenacao:
        base = base.sort_values(ordenacao, kind="stable").reset_index(drop=True)
    base = base.drop(columns=["_ORDEM_STATUS"], errors="ignore")
    with pd.ExcelWriter(FINAL_XLSX_PATH, engine="openpyxl") as writer:
        ufs = sorted(uf for uf in base["UF"].dropna().astype(str).str.strip().str.upper().unique().tolist() if uf)
        if not ufs:
            base.to_excel(writer, sheet_name="SEM_UF", index=False)
            formatar_aba_excel(writer.book["SEM_UF"])
            return
        abas_usadas: set[str] = set()
        for uf in ufs:
            aba = nome_aba_uf(uf)
            while aba in abas_usadas:
                aba = f"{aba[:28]}_{len(abas_usadas) + 1}"
            abas_usadas.add(aba)
            df_uf = base[base["UF"].astype(str).str.upper().eq(uf)].copy()
            df_uf.to_excel(writer, sheet_name=aba, index=False)
            formatar_aba_excel(writer.book[aba])


def main() -> int:
    log("Iniciando consolidacao dos arquivos Mercado Farma")
    BASE_DIR.mkdir(parents=True, exist_ok=True)
    PARCIAIS_DIR.mkdir(parents=True, exist_ok=True)

    status_por_uf = ler_status_ufs()
    frames: list[pd.DataFrame] = []
    ufs_atualizadas: set[str] = set()
    for path in arquivos_parciais():
        try:
            df = normalizar_uf(ler_csv(path))
        except Exception as exc:
            log(f"Aviso: nao consegui ler {path}: {exc}")
            continue

        valido, motivo = csv_valido(df)
        if not valido:
            log(f"Aviso: CSV invalido ignorado: {path} ({motivo})")
            continue

        uf = uf_arquivo(path, df)
        status_uf = status_por_uf.get(uf, {})
        if status_uf and str(status_uf.get("status", "")).lower() != "sucesso":
            log(f"Aviso: CSV ignorado porque a UF {uf} terminou com status {status_uf.get('status')}: {status_uf.get('erro', '')}")
            continue

        if uf:
            ufs_atualizadas.add(uf)
        frames.append(df)
        log(f"Arquivo parcial localizado: {path} ({len(df)} linhas)")

    anterior = pd.DataFrame()
    if FINAL_PATH.exists():
        try:
            anterior = normalizar_uf(ler_csv(FINAL_PATH))
            if ufs_atualizadas and not anterior.empty:
                anterior = anterior[~anterior["UF"].isin(ufs_atualizadas)].copy()
                log("Base anterior carregada para preservar UFs sem nova extracao.")
        except Exception as exc:
            log(f"Aviso: nao consegui carregar consolidado anterior: {exc}")

    if not frames and anterior.empty:
        log("Aviso: nenhuma UF gerou arquivo parcial valido e nao existe consolidado anterior.")
        STATUS_PATH.write_text(
            json.dumps(
                {"gerado_em": agora_brasilia_iso(), "ufs_atualizadas": [], "status": list(status_por_uf.values())},
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return 0

    if frames:
        consolidado = pd.concat([anterior, *frames], ignore_index=True) if not anterior.empty else pd.concat(frames, ignore_index=True)
    else:
        consolidado = anterior.copy()
    subset = [col for col in ["UF", "EAN", "DISTRIBUIDORA", "CNPJ_REFERENCIA"] if col in consolidado.columns]
    if subset:
        consolidado = consolidado.drop_duplicates(subset=subset, keep="last")
    consolidado.to_csv(FINAL_PATH, index=False, encoding="utf-8-sig")
    salvar_excel_por_uf(consolidado)

    resumo = {
        "gerado_em": agora_brasilia_iso(),
        "arquivo": FINAL_PATH.relative_to(ROOT).as_posix(),
        "arquivo_excel": FINAL_XLSX_PATH.relative_to(ROOT).as_posix(),
        "ufs_atualizadas": sorted(ufs_atualizadas),
        "total_linhas": int(len(consolidado)),
        "status": list(status_por_uf.values()),
    }
    STATUS_PATH.write_text(json.dumps(resumo, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"Arquivo consolidado gerado com sucesso: {FINAL_PATH}")
    log(f"Arquivo Excel por UF gerado com sucesso: {FINAL_XLSX_PATH}")
    log(f"Total de linhas consolidadas: {len(consolidado)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
