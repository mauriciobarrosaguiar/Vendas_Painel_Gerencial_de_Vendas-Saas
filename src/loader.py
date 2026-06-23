from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Callable

import pandas as pd
import streamlit as st

from src.configuracoes import aplicar_ajustes_vendedores
from src.datas import agora_brasilia
from src.persistencia import carregar_bytes, carregar_metadados, criar_backup, existe_persistido, restaurar_backup, salvar_bytes
from src.tratamento import (
    COLUNAS_ACOES,
    COLUNAS_BUSSOLA,
    COLUNAS_CONTATO,
    COLUNAS_PAINEL,
    COLUNAS_PRODUTOS_MIX,
    TIPO_SEM_CLASSIFICACAO,
    deduplicar_exportacao_bussola,
    normalizar_ean,
    padronizar_colunas,
    preparar_acoes,
    preparar_base_vendas,
    preparar_painel_equipe,
    preparar_produtos_mix,
    validar_colunas_esperadas,
)


ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"
MERCADO_FARMA_CONSOLIDADO = DATA_DIR / "mercadofarma" / "mercadofarma_consolidado.csv"

BASES_CRITICAS = {"bussola", "painel", "produtos_mix", "mercado_farma", "produtos_mercado_farma"}

NOMES_BASES = {
    "bussola": "Bússola",
    "painel": "Painel de Clientes",
    "produtos_mix": "Produtos/Mix",
    "mercado_farma": "Mercado Farma",
    "produtos_mercado_farma": "Produtos Mercado Farma",
}

ARQUIVOS_PADRAO = {
    "bussola": DATA_DIR / "bussola.xlsx",
    "painel": DATA_DIR / "PAINEL EQUIPE NORTE.xlsx",
    "acoes": DATA_DIR / "template_acoes_promocionais.xlsx",
    "produtos_mix": DATA_DIR / "template_produtos_mix.xlsx",
    "mercado_farma": DATA_DIR / "mercado_farma.xlsx",
    "produtos_mercado_farma": DATA_DIR / "produtos.xlsx",
    "bussola_historico": DATA_DIR / "bussola_historico.xlsx",
}

ABAS_PADRAO = {
    "bussola": "Pedidos",
    "painel": "Planilha1",
    "acoes": 0,
    "produtos_mix": 0,
    "mercado_farma": 0,
    "produtos_mercado_farma": 0,
    "bussola_historico": "Pedidos",
}


def _uploads_sessao() -> dict[str, dict[str, object]]:
    return st.session_state.setdefault("uploads_bases", {})


def _cache_state_get(chave: str, padrao: object = None) -> object:
    try:
        return st.session_state.get(chave, padrao)
    except Exception:
        return padrao


def _cache_state_set(chave: str, valor: object) -> None:
    try:
        st.session_state[chave] = valor
    except Exception:
        pass


def _cache_state_pop(chave: str) -> None:
    try:
        st.session_state.pop(chave, None)
    except Exception:
        pass


def limpar_cache_dados() -> None:
    limpar_cache_dados()


def _versao_cache_global() -> str:
    partes: dict[str, object] = {"metadata": {}, "sessao": {}}
    try:
        metadados = carregar_metadados()
    except Exception:
        metadados = {}
    if isinstance(metadados, dict):
        partes["metadata"] = {
            chave: item.get("updated_at", "")
            for chave, item in metadados.items()
            if isinstance(item, dict)
        }
    try:
        uploads = _uploads_sessao()
    except Exception:
        uploads = {}
    partes["sessao"] = {
        chave: {
            "updated_at": upload.get("updated_at", ""),
            "size": upload.get("size", 0),
        }
        for chave, upload in uploads.items()
        if isinstance(upload, dict)
    }
    if MERCADO_FARMA_CONSOLIDADO.exists():
        try:
            partes["mercadofarma_consolidado_mtime"] = MERCADO_FARMA_CONSOLIDADO.stat().st_mtime
        except OSError:
            pass
    return json.dumps(partes, sort_keys=True, default=str)


def _versao_cache(chave: str) -> str:
    versao_sessao = str(_cache_state_get(f"{chave}_updated_at", "") or "")
    return json.dumps({"global": _versao_cache_global(), "base": chave, "sessao": versao_sessao}, sort_keys=True)


def _df_vazio(df: pd.DataFrame | None) -> bool:
    return df is None or df.empty


def _ler_excel_upload(conteudo: bytes, sheet_name: str | int = 0) -> pd.DataFrame:
    if not conteudo:
        return pd.DataFrame()
    return pd.read_excel(BytesIO(conteudo), sheet_name=sheet_name, dtype=str, engine="openpyxl")


def _formatar_mb(tamanho_bytes: int) -> str:
    return f"{tamanho_bytes / (1024 * 1024):.2f} MB"


def _celula_vazia(valor: object) -> bool:
    if valor is None:
        return True
    if pd.isna(valor):
        return True
    return str(valor).strip().lower() in {"", "nan", "none", "<na>", "nat", "null"}


def _colunas_unicas(cabecalho: list[object]) -> list[str]:
    nomes: list[str] = []
    contagem: dict[str, int] = {}
    for idx, valor in enumerate(cabecalho):
        nome = str(valor).strip() if not _celula_vazia(valor) else f"Unnamed: {idx}"
        ocorrencias = contagem.get(nome, 0)
        contagem[nome] = ocorrencias + 1
        nomes.append(nome if ocorrencias == 0 else f"{nome}.{ocorrencias}")
    return nomes


def _ler_painel_upload(conteudo: bytes) -> pd.DataFrame:
    if not conteudo:
        return pd.DataFrame()
    try:
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb[ABAS_PADRAO["painel"]] if ABAS_PADRAO["painel"] in wb.sheetnames else wb.worksheets[0]
        cabecalho: list[str] | None = None
        linhas: list[list[object]] = []
        vazias_consecutivas = 0
        for linha in ws.iter_rows(values_only=True):
            valores = list(linha)
            if all(_celula_vazia(valor) for valor in valores):
                if cabecalho is not None:
                    vazias_consecutivas += 1
                    if vazias_consecutivas >= 500:
                        break
                continue
            vazias_consecutivas = 0
            if cabecalho is None:
                cabecalho = _colunas_unicas(valores)
                continue
            if len(valores) < len(cabecalho):
                valores.extend([""] * (len(cabecalho) - len(valores)))
            linhas.append(valores[: len(cabecalho)])
        wb.close()
        return pd.DataFrame(linhas, columns=cabecalho or [])
    except Exception:
        try:
            return _ler_excel_upload(conteudo, ABAS_PADRAO["painel"])
        except ValueError:
            return _ler_excel_upload(conteudo, 0)


def _cnpj_valido(cnpj: object) -> bool:
    texto = str(cnpj or "").strip()
    return bool(texto and texto.isdigit() and len(texto) == 14 and texto != "0" * 14 and len(set(texto)) > 1)


def _linha_util(serie: pd.Series) -> bool:
    for valor in serie:
        if not _celula_vazia(valor):
            return True
    return False


def _compactar_upload_painel(conteudo: bytes) -> tuple[bytes, int]:
    painel_raw = _ler_painel_upload(conteudo)
    painel = preparar_painel_equipe(painel_raw)
    colunas_compactas_padrao = COLUNAS_PAINEL + COLUNAS_CONTATO + ["cnpj_limpo", "grupo_sip", "cliente_ativo"]
    if painel.empty:
        compacto = pd.DataFrame(columns=colunas_compactas_padrao)
    else:
        colunas_usuario = [col for col in COLUNAS_PAINEL + COLUNAS_CONTATO if col in painel.columns]
        if colunas_usuario:
            painel = painel[painel[colunas_usuario].apply(_linha_util, axis=1)]
        painel = painel[painel["cnpj_limpo"].apply(_cnpj_valido)]
        painel = painel.drop_duplicates("cnpj_limpo", keep="first").reset_index(drop=True)
        colunas_compactas = [col for col in colunas_compactas_padrao if col in painel.columns]
        compacto = painel[colunas_compactas].copy()

    saida = BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        compacto.to_excel(writer, sheet_name="Planilha1", index=False)
    return saida.getvalue(), int(len(compacto.index))


def _tem_coluna(df: pd.DataFrame, nomes: list[str]) -> bool:
    if df is None or df.empty:
        return False
    colunas = set(padronizar_colunas(df).columns)
    return any(nome in colunas for nome in nomes)


def _validar_upload_produtos_mix(conteudo: bytes) -> tuple[bool, str]:
    try:
        bruto = _ler_excel_upload(conteudo, ABAS_PADRAO["produtos_mix"])
    except Exception as exc:
        return False, f"Nao consegui ler o arquivo: {exc}"

    if bruto.empty:
        return False, "O arquivo esta vazio."
    if not _tem_coluna(bruto, ["ean"]):
        return False, "A coluna EAN nao foi encontrada."
    if not _tem_coluna(bruto, ["produto", "principio_ativo", "nome_do_produto", "descricao"]):
        return False, "A coluna Produto nao foi encontrada."
    if not _tem_coluna(bruto, ["tipo_mix", "tipo", "mix", "classificacao", "categoria"]):
        return False, "A coluna Tipo Mix nao foi encontrada."

    tratado = preparar_produtos_mix(bruto)
    eans_validos = int(tratado["ean_limpo"].dropna().astype(str).str.strip().ne("").sum()) if "ean_limpo" in tratado else 0
    classificados = tratado[tratado["tipo_mix"].ne(TIPO_SEM_CLASSIFICACAO)] if "tipo_mix" in tratado else pd.DataFrame()
    if eans_validos < 10:
        return False, "A planilha precisa ter pelo menos 10 EANs validos."
    if classificados.empty:
        return False, "Todos os produtos ficaram SEM CLASSIFICACAO."
    return True, ""


def _validar_upload_produtos_mercado_farma(conteudo: bytes) -> tuple[bool, str]:
    try:
        bruto = _ler_excel_upload(conteudo, ABAS_PADRAO["produtos_mercado_farma"])
    except Exception as exc:
        return False, f"Nao consegui ler a planilha produtos.xlsx: {exc}"
    if bruto.empty:
        return False, "A planilha produtos.xlsx esta vazia."
    base = padronizar_colunas(bruto)
    coluna = "ean" if "ean" in base.columns else base.columns[0] if len(base.columns) else ""
    eans = base[coluna].dropna().astype(str).map(normalizar_ean) if coluna else pd.Series(dtype=str)
    total = int(eans[eans.ne("")].nunique())
    if total <= 0:
        return False, "A planilha precisa conter EANs validos."
    return True, f"{total} EANs validos encontrados."


def _validar_upload_generico(chave: str, conteudo: bytes) -> tuple[bool, str]:
    if chave == "produtos_mix":
        return _validar_upload_produtos_mix(conteudo)
    if chave == "produtos_mercado_farma":
        return _validar_upload_produtos_mercado_farma(conteudo)
    try:
        bruto = _ler_painel_upload(conteudo) if chave == "painel" else _ler_excel_upload(conteudo, ABAS_PADRAO.get(chave, 0))
    except Exception as exc:
        return False, f"Nao consegui ler o arquivo enviado: {exc}"
    if bruto.empty:
        return False, "O arquivo enviado esta vazio."

    if chave in {"bussola", "bussola_historico"}:
        minimas = ["cnpj_pdv", "ean", "produto", "status_pedido", "pedido_id", "data_do_pedido", "preco_unitario_sem_imposto", "valor_faturado"]
        faltantes = [coluna for coluna in minimas if not _tem_coluna(bruto, [coluna])]
        if faltantes:
            return False, "A base Bussola precisa conter: " + ", ".join(minimas)
        if not (_tem_coluna(bruto, ["quantidade_atendida"]) or _tem_coluna(bruto, ["quantidade_faturada"])):
            return False, "A base Bussola precisa conter quantidade_atendida ou quantidade_faturada."
    elif chave == "painel":
        if not _tem_coluna(bruto, ["cnpj"]):
            return False, "A base de clientes precisa conter CNPJ."
        if not _tem_coluna(bruto, ["nome_pdv", "cliente", "razao_social", "nome"]):
            return False, "A base de clientes precisa conter Nome PDV."
        if not _tem_coluna(bruto, ["cidade"]):
            return False, "A base de clientes precisa conter Cidade."
        if not _tem_coluna(bruto, ["uf"]):
            return False, "A base de clientes precisa conter UF."
        if not _tem_coluna(bruto, ["nome_rep", "consultor", "representante"]):
            return False, "A base de clientes precisa conter Nome REP ou Consultor."
    elif chave == "acoes":
        if not _tem_coluna(bruto, ["campanha", "nome_acao", "tipo_acao"]):
            return False, "A base de acoes precisa conter campanha."
        if not (_tem_coluna(bruto, ["produto"]) or _tem_coluna(bruto, ["ean"])):
            return False, "A base de acoes precisa conter Produto ou EAN."
        if not (_tem_coluna(bruto, ["desconto"]) or _tem_coluna(bruto, ["data_inicio"]) or _tem_coluna(bruto, ["data_fim"])):
            return False, "A base de acoes precisa conter desconto ou validade."
    elif chave == "mercado_farma":
        if not _tem_coluna(bruto, ["ean"]):
            return False, "A base Mercado Farma precisa conter EAN."
        if not _tem_coluna(bruto, ["produto", "nome_do_produto"]):
            return False, "A base Mercado Farma precisa conter Produto."
        if not _tem_coluna(bruto, ["distribuidora"]):
            return False, "A base Mercado Farma precisa conter Distribuidora."
        if not _tem_coluna(bruto, ["preco_sem_imposto", "preco_final", "preco_final_r", "estoque"]):
            return False, "A base Mercado Farma precisa conter preco ou estoque."
    return True, ""


def _validar_base_bussola(df: pd.DataFrame) -> tuple[bool, str]:
    if _df_vazio(df):
        return False, "arquivo vazio"
    if not _tem_coluna(df, ["cnpj_pdv", "cnpj"]):
        return False, "CNPJ não encontrado"
    if not _tem_coluna(df, ["ean"]):
        return False, "EAN não encontrado"
    if not _tem_coluna(df, ["status_pedido", "status"]):
        return False, "status não encontrado"
    if not (
        _tem_coluna(df, ["valor_faturado", "valor_total_solicitado_sem_imposto", "total_atendido_sem_imposto"])
        or _tem_coluna(df, ["quantidade_atendida", "quantidade_faturada", "quantidade_solicitada"])
    ):
        return False, "valor ou quantidade não encontrados"
    return True, ""


def _validar_base_painel(df: pd.DataFrame) -> tuple[bool, str]:
    if _df_vazio(df):
        return False, "arquivo vazio"
    checks = [
        ("CNPJ", ["cnpj"]),
        ("nome PDV", ["nome_pdv", "nome_fantasia", "razao_social", "nome"]),
        ("cidade", ["cidade"]),
        ("UF", ["uf"]),
        ("consultor", ["nome_rep", "consultor", "representante", "nome_consultor_territ"]),
    ]
    faltantes = [nome for nome, aliases in checks if not _tem_coluna(df, aliases)]
    if faltantes:
        return False, "colunas ausentes: " + ", ".join(faltantes)
    return True, ""


def _validar_base_produtos_mix(df: pd.DataFrame) -> tuple[bool, str]:
    if _df_vazio(df):
        return False, "arquivo vazio"
    checks = [
        ("EAN", ["ean"]),
        ("produto", ["produto", "principio_ativo", "nome_do_produto", "descricao"]),
        ("tipo/classificação", ["tipo_mix", "tipo", "mix", "classificacao", "categoria"]),
    ]
    faltantes = [nome for nome, aliases in checks if not _tem_coluna(df, aliases)]
    if faltantes:
        return False, "colunas ausentes: " + ", ".join(faltantes)
    return True, ""


def _validar_base_produtos_mercado_farma(df: pd.DataFrame) -> tuple[bool, str]:
    if _df_vazio(df):
        return False, "arquivo vazio"
    if not _tem_coluna(df, ["ean"]):
        return False, "EAN não encontrado"
    base = padronizar_colunas(df)
    coluna = "ean" if "ean" in base.columns else base.columns[0] if len(base.columns) else ""
    eans = base[coluna].dropna().astype(str).map(normalizar_ean) if coluna else pd.Series(dtype=str)
    if int(eans[eans.ne("")].nunique()) <= 0:
        return False, "nenhum EAN válido encontrado"
    return True, ""


def _validar_base_mercado_farma(df: pd.DataFrame) -> tuple[bool, str]:
    if _df_vazio(df):
        return False, "arquivo vazio"
    if not _tem_coluna(df, ["ean"]):
        return False, "EAN não encontrado"
    faltantes = []
    if not _tem_coluna(df, ["produto", "nome_do_produto", "principio_ativo"]):
        faltantes.append("produto")
    if not _tem_coluna(df, ["distribuidora"]):
        faltantes.append("distribuidora")
    if not _tem_coluna(df, ["preco_sem_imposto", "sem_imposto", "sem_imposto_r", "preco_final", "preco_final_r", "preco_com_imposto", "estoque"]):
        faltantes.append("preço/estoque")
    if faltantes:
        return False, "colunas ausentes: " + ", ".join(faltantes)
    return True, ""


VALIDADORES_CRITICOS: dict[str, Callable[[pd.DataFrame], tuple[bool, str]]] = {
    "bussola": _validar_base_bussola,
    "painel": _validar_base_painel,
    "produtos_mix": _validar_base_produtos_mix,
    "mercado_farma": _validar_base_mercado_farma,
    "produtos_mercado_farma": _validar_base_produtos_mercado_farma,
}


def _salvar_upload(chave: str, arquivo, conteudo: bytes, mensagem_backup: str | None = None) -> bool:
    tamanho_original = len(conteudo)
    clientes_validos: int | None = None
    if chave in {"bussola", "bussola_historico"}:
        bruto = _ler_excel_upload(conteudo, ABAS_PADRAO[chave])
        deduplicado = deduplicar_exportacao_bussola(bruto)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            deduplicado.to_excel(writer, sheet_name=ABAS_PADRAO[chave], index=False)
        conteudo = buffer.getvalue()
    elif chave == "painel":
        conteudo, clientes_validos = _compactar_upload_painel(conteudo)

    atualizado_em = agora_brasilia().isoformat()
    if chave in {"produtos_mix", "produtos_mercado_farma", "bussola", "painel", "mercado_farma", "bussola_historico", "acoes"}:
        try:
            criar_backup(chave, mensagem_backup or f"Backup automatico antes de atualizar {chave}")
        except Exception:
            st.warning("Não foi possível criar backup automático, mas vou tentar salvar a nova base.")
    st.session_state[f"{chave}_updated_at"] = atualizado_em
    st.session_state[f"{chave}_uploaded_name"] = arquivo.name
    _uploads_sessao()[chave] = {
        "name": arquivo.name,
        "bytes": conteudo,
        "updated_at": atualizado_em,
        "size": len(conteudo),
    }
    salvar_bytes(chave, conteudo, f"Atualiza base {chave} pelo painel")
    if chave == "painel" and clientes_validos is not None:
        st.success(
            (
                f"Base de clientes tratada e salva: {clientes_validos:,} clientes válidos. "
                f"Tamanho reduzido de {_formatar_mb(tamanho_original)} para {_formatar_mb(len(conteudo))}."
            ).replace(",", ".")
        )
    st.cache_data.clear()
    return True


def registrar_upload(chave: str, arquivo) -> bool:
    if arquivo is None:
        return False
    conteudo = arquivo.getvalue()
    valido, mensagem = _validar_upload_generico(chave, conteudo)
    if not valido:
        st.error(f"Arquivo invalido para {chave}. A base anterior foi preservada. {mensagem}")
        return False
    if mensagem:
        st.info(mensagem)
    return _salvar_upload(chave, arquivo, conteudo)


def registrar_upload_produtos_mix(arquivo) -> bool:
    if arquivo is None:
        return False
    conteudo = arquivo.getvalue()
    valido, mensagem = _validar_upload_produtos_mix(conteudo)
    if not valido:
        st.error(
            "Arquivo de Produtos / Mix invalido. A base anterior foi preservada. "
            "Envie a planilha correta com EAN, Produto e Tipo Mix. "
            f"Detalhe: {mensagem}"
        )
        return False
    return _salvar_upload("produtos_mix", arquivo, conteudo, "Backup automatico de Produtos / Mix")


def registrar_upload_produtos_mercado_farma(arquivo) -> bool:
    if arquivo is None:
        return False
    conteudo = arquivo.getvalue()
    valido, mensagem = _validar_upload_produtos_mercado_farma(conteudo)
    if not valido:
        st.error(f"Arquivo produtos.xlsx invalido. A lista anterior foi preservada. {mensagem}")
        return False
    st.info(mensagem)
    return _salvar_upload("produtos_mercado_farma", arquivo, conteudo, "Backup automatico de produtos Mercado Farma")


def restaurar_backup_produtos_mix() -> bool:
    ok = restaurar_backup("produtos_mix")
    if ok:
        _uploads_sessao().pop("produtos_mix", None)
        st.session_state.pop("produtos_mix_updated_at", None)
        st.session_state.pop("produtos_mix_uploaded_name", None)
        limpar_cache_dados()
    return ok


def limpar_uploads() -> None:
    st.session_state["uploads_bases"] = {}
    limpar_cache_dados()


def fonte_ativa(chave: str) -> str:
    upload = _uploads_sessao().get(chave)
    if upload:
        return f"Upload salvo: {upload.get('name', '')}"
    if chave == "mercado_farma" and MERCADO_FARMA_CONSOLIDADO.exists():
        return "Consolidado GitHub Actions"
    if existe_persistido(chave):
        return "Base salva"
    caminho = ARQUIVOS_PADRAO[chave]
    return f"Pasta data: {caminho.name}" if caminho.exists() else "Arquivo não encontrado"


@st.cache_data(show_spinner=False)
def _ler_excel_bytes(conteudo: bytes, sheet_name: str | int, versao_cache: str = "") -> pd.DataFrame:
    return pd.read_excel(BytesIO(conteudo), sheet_name=sheet_name, dtype=str, engine="openpyxl")


@st.cache_data(show_spinner=False)
def _ler_excel_caminho(caminho: str, sheet_name: str | int, mtime: float, versao_cache: str = "") -> pd.DataFrame:
    return pd.read_excel(caminho, sheet_name=sheet_name, dtype=str, engine="openpyxl")


@st.cache_data(show_spinner=False)
def _ler_painel_bytes(conteudo: bytes, versao_cache: str = "") -> pd.DataFrame:
    return _ler_painel_upload(conteudo)


@st.cache_data(show_spinner=False)
def _ler_painel_caminho(caminho: str, mtime: float, versao_cache: str = "") -> pd.DataFrame:
    return _ler_painel_upload(Path(caminho).read_bytes())


@st.cache_data(show_spinner=False)
def _ler_csv_caminho(caminho: str, mtime: float, versao_cache: str = "") -> pd.DataFrame:
    return pd.read_csv(caminho, dtype=str, sep=None, engine="python")


def _ler_base_fonte(chave: str, origem: str, valor: object) -> pd.DataFrame:
    versao = _versao_cache(chave)
    if origem == "sessao":
        return valor.copy() if isinstance(valor, pd.DataFrame) else pd.DataFrame()
    if origem == "consolidado":
        caminho = Path(str(valor))
        return _ler_csv_caminho(str(caminho), caminho.stat().st_mtime, versao)
    if origem == "arquivo":
        caminho = Path(str(valor))
        if chave == "painel":
            return _ler_painel_caminho(str(caminho), caminho.stat().st_mtime, versao)
        return _ler_excel_caminho(str(caminho), ABAS_PADRAO[chave], caminho.stat().st_mtime, versao)
    if isinstance(valor, bytes):
        if chave == "painel":
            return _ler_painel_bytes(valor, versao)
        return _ler_excel_bytes(valor, ABAS_PADRAO[chave], versao)
    return pd.DataFrame()


def _base_sessao_key(chave: str) -> str:
    return f"_ultima_base_valida_{chave}"


def _guardar_base_valida(chave: str, df: pd.DataFrame, origem: str) -> None:
    if _df_vazio(df):
        return
    _cache_state_set(_base_sessao_key(chave), df.copy())
    _cache_state_set(f"{_base_sessao_key(chave)}_origem", origem)


def _ultima_base_valida(chave: str) -> pd.DataFrame | None:
    base = _cache_state_get(_base_sessao_key(chave))
    return base.copy() if isinstance(base, pd.DataFrame) and not base.empty else None


def _carregar_bytes_seguro(chave: str) -> bytes | None:
    try:
        return carregar_bytes(chave)
    except Exception as exc:
        _cache_state_set(f"_falha_leitura_{chave}", str(exc))
        return None


def _candidatos_base(chave: str) -> list[tuple[str, object]]:
    candidatos: list[tuple[str, object]] = []
    upload = _uploads_sessao().get(chave)
    if upload and upload.get("bytes"):
        candidatos.append(("upload da sessão", upload["bytes"]))

    if chave == "mercado_farma" and MERCADO_FARMA_CONSOLIDADO.exists():
        candidatos.append(("consolidado", MERCADO_FARMA_CONSOLIDADO))

    persistido = _carregar_bytes_seguro(chave)
    if persistido:
        candidatos.append(("persistência", persistido))

    caminho = ARQUIVOS_PADRAO.get(chave)
    if caminho and caminho.exists():
        candidatos.append(("arquivo", caminho))

    chave_backup = f"{chave}_backup"
    backup = _carregar_bytes_seguro(chave_backup)
    if backup:
        candidatos.append(("backup", backup))

    ultima = _ultima_base_valida(chave)
    if ultima is not None:
        candidatos.append(("sessao", ultima))
    return candidatos


def _warning_base_critica(chave: str, mensagem: str) -> None:
    if chave not in BASES_CRITICAS:
        return
    try:
        st.warning(mensagem)
    except Exception:
        pass


def obter_base_segura(
    chave: str,
    leitor: Callable[[str, str, object], pd.DataFrame],
    validador: Callable[[pd.DataFrame], tuple[bool, str]],
) -> pd.DataFrame:
    """Lê uma base crítica com validação e fallback sem trocar base boa por vazio."""
    nome = NOMES_BASES.get(chave, chave)
    falhas: list[str] = []
    candidatos = _candidatos_base(chave)
    falha_leitura = _cache_state_get(f"_falha_leitura_{chave}", "")
    if falha_leitura:
        falhas.append(f"persistência: erro de leitura ({falha_leitura})")
        _cache_state_pop(f"_falha_leitura_{chave}")

    for origem, valor in candidatos:
        try:
            df = leitor(chave, origem, valor)
        except Exception as exc:
            falhas.append(f"{origem}: erro de leitura ({exc})")
            continue

        valido, motivo = validador(df)
        if not valido:
            falhas.append(f"{origem}: {motivo}")
            continue

        if falhas and origem != "upload da sessão":
            _warning_base_critica(
                chave,
                f"Base {nome} não carregou corretamente. Usando última base válida/backup.",
            )
        _guardar_base_valida(chave, df, origem)
        return df

    ultima = _ultima_base_valida(chave)
    if ultima is not None:
        _warning_base_critica(
            chave,
            f"Base {nome} não carregou corretamente. Usando última base válida/backup.",
        )
        return ultima

    detalhe = "; ".join(falhas[-3:]) if falhas else "nenhuma fonte encontrada"
    _warning_base_critica(
        chave,
        f"Base {nome} não carregou corretamente. Não encontrei última base válida/backup para usar. Detalhe: {detalhe}",
    )
    return pd.DataFrame()


def _carregar_excel(chave: str) -> pd.DataFrame:
    if chave in VALIDADORES_CRITICOS:
        return obter_base_segura(chave, _ler_base_fonte, VALIDADORES_CRITICOS[chave])

    upload = _uploads_sessao().get(chave)
    if upload and upload.get("bytes"):
        return _ler_excel_bytes(upload["bytes"], ABAS_PADRAO[chave], _versao_cache(chave))

    persistido = carregar_bytes(chave)
    if persistido:
        return _ler_excel_bytes(persistido, ABAS_PADRAO[chave], _versao_cache(chave))

    caminho = ARQUIVOS_PADRAO[chave]
    if not caminho.exists():
        return pd.DataFrame()
    return _ler_excel_caminho(str(caminho), ABAS_PADRAO[chave], caminho.stat().st_mtime, _versao_cache(chave))


def carregar_bussola() -> pd.DataFrame:
    return _carregar_excel("bussola")


def carregar_bussola_historico() -> pd.DataFrame:
    return _carregar_excel("bussola_historico")


def carregar_painel_equipe() -> pd.DataFrame:
    return _carregar_excel("painel")


def carregar_acoes() -> pd.DataFrame:
    return _carregar_excel("acoes")


def carregar_produtos_mix() -> pd.DataFrame:
    return _carregar_excel("produtos_mix")


def carregar_mercado_farma() -> pd.DataFrame:
    return obter_base_segura("mercado_farma", _ler_base_fonte, _validar_base_mercado_farma)


def carregar_produtos_mercado_farma() -> pd.DataFrame:
    return _carregar_excel("produtos_mercado_farma")


def _base_tratada_key(chave: str) -> str:
    return f"_ultima_base_tratada_valida_{chave}"


def _guardar_base_tratada(chave: str, df: pd.DataFrame) -> None:
    if not _df_vazio(df):
        _cache_state_set(_base_tratada_key(chave), df.copy())


def _ultima_base_tratada(chave: str) -> pd.DataFrame | None:
    base = _cache_state_get(_base_tratada_key(chave))
    return base.copy() if isinstance(base, pd.DataFrame) and not base.empty else None


def _base_tratada_segura(chave: str, df: pd.DataFrame, nome_base: str) -> pd.DataFrame:
    if not _df_vazio(df):
        _guardar_base_tratada(chave, df)
        return df
    anterior = _ultima_base_tratada(chave)
    if anterior is not None:
        try:
            st.warning(f"Base {nome_base} não carregou corretamente. Usando última base válida/backup.")
        except Exception:
            pass
        return anterior
    if chave in {"clientes", "vendas", "produtos_mix"}:
        try:
            st.warning(f"Base {nome_base} não carregou corretamente. Não encontrei última base válida/backup para usar.")
        except Exception:
            pass
    return df


def proteger_dados_pagina(dados: dict[str, object], pagina: str = "") -> tuple[pd.DataFrame, pd.DataFrame]:
    vendas = dados.get("vendas") if isinstance(dados, dict) else pd.DataFrame()
    clientes = dados.get("clientes") if isinstance(dados, dict) else pd.DataFrame()
    vendas = vendas if isinstance(vendas, pd.DataFrame) else pd.DataFrame()
    clientes = clientes if isinstance(clientes, pd.DataFrame) else pd.DataFrame()

    if vendas.empty:
        anterior_vendas = _ultima_base_tratada("vendas")
        if anterior_vendas is not None:
            st.warning("Base Bússola não carregou corretamente. Usando última base válida/backup.")
            vendas = anterior_vendas

    if clientes.empty:
        anterior_clientes = _ultima_base_tratada("clientes")
        if anterior_clientes is not None:
            st.warning("Base Painel de Clientes não carregou corretamente. Usando última base válida/backup.")
            clientes = anterior_clientes

    if clientes.empty:
        st.error(
            "A base de clientes não carregou corretamente. Para evitar indicadores zerados, "
            "a página foi interrompida. Recarregue a persistência ou restaure um backup em Importação."
        )
        st.stop()

    if vendas.empty:
        st.error(
            "A base Bússola não carregou corretamente e não há uma base anterior válida na sessão. "
            "Para evitar indicadores zerados, a página foi interrompida."
        )
        st.stop()

    _guardar_base_tratada("vendas", vendas)
    _guardar_base_tratada("clientes", clientes)
    return vendas, clientes


def carregar_dados_tratados() -> dict[str, pd.DataFrame | list[str]]:
    bussola_atual_raw = carregar_bussola()
    bussola_historico_raw = carregar_bussola_historico()
    bussola_raw = bussola_atual_raw
    painel_raw = carregar_painel_equipe()
    acoes_raw = carregar_acoes()
    produtos_raw = carregar_produtos_mix()
    mercado_raw = carregar_mercado_farma()
    produtos_mercado_raw = carregar_produtos_mercado_farma()

    avisos: list[str] = []
    avisos.extend(validar_colunas_esperadas(bussola_raw, COLUNAS_BUSSOLA, "bussola.xlsx"))
    avisos.extend(validar_colunas_esperadas(painel_raw, COLUNAS_PAINEL, "PAINEL EQUIPE NORTE.xlsx"))
    if acoes_raw.empty:
        avisos.append("template_acoes_promocionais.xlsx: sem ações cadastradas. Use a tela Importar Bases para baixar o modelo.")
    if produtos_raw.empty:
        avisos.append("template_produtos_mix.xlsx: sem produtos classificados. Produtos vendidos ficarão como SEM CLASSIFICACAO.")
    else:
        avisos.extend(validar_colunas_esperadas(produtos_raw, COLUNAS_PRODUTOS_MIX, "template_produtos_mix.xlsx"))
    if not acoes_raw.empty:
        avisos.extend(validar_colunas_esperadas(acoes_raw, COLUNAS_ACOES, "template_acoes_promocionais.xlsx"))

    clientes = aplicar_ajustes_vendedores(preparar_painel_equipe(painel_raw))
    produtos_mix = preparar_produtos_mix(produtos_raw)
    acoes = preparar_acoes(acoes_raw)
    clientes = _base_tratada_segura("clientes", clientes, "Painel de Clientes")
    produtos_mix = _base_tratada_segura("produtos_mix", produtos_mix, "Produtos/Mix")
    vendas = preparar_base_vendas(bussola_raw, clientes, produtos_mix)
    vendas = _base_tratada_segura("vendas", vendas, "Bússola")

    return {
        "vendas": vendas,
        "clientes": clientes,
        "produtos_mix": produtos_mix,
        "acoes": acoes,
        "avisos": avisos,
        "raw_bussola": bussola_atual_raw,
        "raw_bussola_historico": bussola_historico_raw,
        "raw_bussola_completa": bussola_raw,
        "raw_painel": painel_raw,
        "raw_acoes": acoes_raw,
        "raw_produtos_mix": produtos_raw,
        "mercado_farma": mercado_raw,
        "produtos_mercado_farma": produtos_mercado_raw,
    }


def modelo_acoes() -> pd.DataFrame:
    return pd.DataFrame(columns=COLUNAS_ACOES)


def modelo_produtos_mix() -> pd.DataFrame:
    return pd.DataFrame(columns=COLUNAS_PRODUTOS_MIX)
