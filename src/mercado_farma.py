from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path
import shutil
import threading
from typing import Callable
from uuid import uuid4

import numpy as np
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

from src.datas import agora_brasilia
from src.loader import DATA_DIR, carregar_mercado_farma
from src.mercadofarma_inventory import login_mercadofarma, processar_ean_catalogo, selecionar_cnpj_catalogo
from src.persistencia import carregar_json, salvar_bytes, salvar_json
from src.tratamento import converter_numero, formatar_moeda, normalizar_cnpj, normalizar_ean, padronizar_colunas


COLUNAS_MERCADO = [
    "consultor",
    "uf",
    "cnpj_referencia",
    "ean",
    "produto",
    "distribuidora",
    "estoque",
    "desconto",
    "pf_dist",
    "pf_fabrica",
    "preco_com_imposto",
    "preco_sem_imposto",
    "data_atualizacao",
    "status",
    "erro",
]

JOB_KEY = "mercado_farma_job"
JOB_PARTIAL_PATH = DATA_DIR / "_mercado_farma_parcial.xlsx"
MERCADO_FARMA_DIR = DATA_DIR / "mercadofarma"
MERCADO_FARMA_CONSOLIDADO = MERCADO_FARMA_DIR / "mercadofarma_consolidado.csv"
MERCADO_FARMA_STATUS = MERCADO_FARMA_DIR / "status_mercadofarma.json"
DESCONTOS_KEY = "mercado_farma_descontos"
FLUSH_EVERY_EANS = 25
_THREADS: dict[str, threading.Thread] = {}
_CANCEL_FLAGS: dict[str, threading.Event] = {}

VALID_UFS = {
    "AC",
    "AL",
    "AP",
    "AM",
    "BA",
    "CE",
    "DF",
    "ES",
    "GO",
    "MA",
    "MT",
    "MS",
    "MG",
    "PA",
    "PB",
    "PR",
    "PE",
    "PI",
    "RJ",
    "RN",
    "RS",
    "RO",
    "RR",
    "SC",
    "SP",
    "SE",
    "TO",
}


def _texto(valor: object) -> str:
    return "" if valor is None or pd.isna(valor) else str(valor).strip()


def _norm_nome(valor: object) -> str:
    return " ".join(_texto(valor).upper().split())


def _secret(nome: str, padrao: str = "") -> str:
    try:
        import streamlit as st

        if nome in st.secrets:
            return str(st.secrets[nome])
    except Exception:
        pass
    return str(os.environ.get(nome, padrao) or padrao)


def mascarar_usuario(usuario: object) -> str:
    texto = _texto(usuario)
    if not texto:
        return ""
    if "@" in texto:
        nome, dominio = texto.split("@", 1)
        if len(nome) <= 2:
            nome_mask = nome[:1] + "***"
        else:
            nome_mask = nome[:2] + "***" + nome[-1:]
        return f"{nome_mask}@{dominio}"
    if len(texto) <= 4:
        return texto[:1] + "***"
    return texto[:2] + "***" + texto[-2:]


def carregar_credenciais_mercadofarma(login: dict | None = None, *, exigir: bool = False) -> dict[str, object]:
    fontes = []
    usuario = _secret("MERCADOFARMA_USUARIO")
    senha = _secret("MERCADOFARMA_SENHA")
    if usuario or senha:
        fontes.append("secrets")

    login = login if isinstance(login, dict) else {}
    candidatos = [
        ("configuracao_mercadofarma", login.get("mercadofarma", {})),
        ("configuracao_mercado_farma", login.get("mercado_farma", {})),
        ("configuracao_gd", login.get("gd", {})),
    ]
    for fonte, dados in candidatos:
        if usuario and senha:
            break
        if not isinstance(dados, dict):
            continue
        usuario = usuario or _texto(dados.get("usuario"))
        senha = senha or _texto(dados.get("senha"))
        if usuario or senha:
            fontes.append(fonte)

    faltantes = []
    if not usuario:
        faltantes.append("MERCADOFARMA_USUARIO")
    if not senha:
        faltantes.append("MERCADOFARMA_SENHA")
    configurado = not faltantes
    if exigir and not configurado:
        raise RuntimeError(
            "Configure o acesso GD do Mercado Farma. Secrets ausentes: " + ", ".join(faltantes)
        )
    return {
        "usuario": usuario,
        "senha": senha,
        "usuario_mascarado": mascarar_usuario(usuario),
        "configurado": configurado,
        "faltantes": faltantes,
        "fonte": fontes[0] if fontes else "",
    }


def _credenciais_gd_da_lista(credenciais: list[dict[str, str]] | None) -> tuple[str, str]:
    for item in credenciais or []:
        consultor = _norm_nome(item.get("consultor"))
        if consultor in {"GD", "GERENTE DISTRITAL", "MERCADO FARMA GD"}:
            usuario = _texto(item.get("usuario"))
            senha = _texto(item.get("senha"))
            if usuario and senha:
                return usuario, senha
    return "", ""


def _estado_padrao() -> dict:
    return {
        "job_id": "",
        "status": "parado",
        "mensagem": "",
        "erro": "",
        "logs": [],
        "inicio": "",
        "fim": "",
        "total_eans": 0,
        "total_alvos": 0,
        "target_index": 0,
        "ean_index": 0,
        "processados": 0,
        "total_passos": 0,
        "current_uf": "",
        "current_consultor": "",
        "current_cnpj": "",
        "current_ean": "",
        "alvos": [],
        "cancelar": False,
        "partial_path": str(JOB_PARTIAL_PATH),
    }


def carregar_estado_extracao() -> dict:
    estado = carregar_json(JOB_KEY, _estado_padrao())
    if not isinstance(estado, dict):
        estado = _estado_padrao()
    base = _estado_padrao()
    base.update(estado)
    job_id = str(base.get("job_id", ""))
    thread = _THREADS.get(job_id)
    if base.get("status") == "rodando" and thread is not None and thread.is_alive():
        base["thread_alive"] = True
    else:
        base["thread_alive"] = False
        if base.get("status") == "rodando":
            base["status"] = "interrompido"
            base["mensagem"] = "A extração não está mais em execução. Você pode retomar de onde parou."
            _salvar_estado_extracao(base)
    return base


def extracao_em_execucao() -> bool:
    estado = carregar_estado_extracao()
    job_id = str(estado.get("job_id", ""))
    thread = _THREADS.get(job_id)
    return bool(thread is not None and thread.is_alive())


def _salvar_estado_extracao(estado: dict) -> None:
    salvar_json(JOB_KEY, estado, "Atualiza status da extração Mercado Farma")


def _log_estado(estado: dict, mensagem: str) -> None:
    logs = list(estado.get("logs", []))
    logs.append(f"{agora_brasilia().strftime('%d/%m/%Y %H:%M:%S')} - {mensagem}")
    estado["logs"] = logs[-80:]
    estado["mensagem"] = mensagem


def limpar_estado_extracao() -> dict:
    estado = _estado_padrao()
    estado["mensagem"] = "Pronto para iniciar uma nova extração."
    _salvar_estado_extracao(estado)
    return estado


def ufs_validas_clientes(clientes: pd.DataFrame) -> list[str]:
    if clientes is None or clientes.empty or "uf" not in clientes.columns:
        return []
    base = clientes.copy()
    if "cliente_ativo" in base.columns:
        base = base[base["cliente_ativo"].fillna(True)].copy()
    ufs = base["uf"].dropna().astype(str).str.strip().str.upper()
    return sorted(uf for uf in ufs.unique().tolist() if uf in VALID_UFS)


def preparar_mercado_farma(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=COLUNAS_MERCADO)
    base = df.copy()
    if all(str(col).startswith("Unnamed") for col in base.columns) and len(base) > 0:
        base.columns = base.iloc[0].tolist()
        base = base.iloc[1:].copy()
    base = padronizar_colunas(base)
    aliases = {
        "cnpj": "cnpj_referencia",
        "cnpj_ref": "cnpj_referencia",
        "cnpj_referencia": "cnpj_referencia",
        "ean": "ean",
        "nome_do_produto": "produto",
        "produto": "produto",
        "principio_ativo": "produto",
        "distribuidora": "distribuidora",
        "estoque": "estoque",
        "desconto": "desconto",
        "desconto_percent": "desconto",
        "desconto_percentual": "desconto",
        "pf_dist_r": "pf_dist",
        "pf_dist": "pf_dist",
        "pf_fabrica_r": "pf_fabrica",
        "pf_fabrica": "pf_fabrica",
        "preco_final_r": "preco_com_imposto",
        "preco_final": "preco_com_imposto",
        "preco_com_imposto": "preco_com_imposto",
        "sem_imposto_r": "preco_sem_imposto",
        "sem_imposto": "preco_sem_imposto",
        "preco_sem_imposto": "preco_sem_imposto",
        "data": "data_atualizacao",
        "data_atualizacao": "data_atualizacao",
        "status": "status",
        "erro": "erro",
        "uf": "uf",
        "consultor": "consultor",
        "consultor_usado": "consultor",
    }
    for origem, destino in aliases.items():
        if origem in base.columns and destino not in base.columns:
            base = base.rename(columns={origem: destino})
    for coluna in COLUNAS_MERCADO:
        if coluna not in base.columns:
            base[coluna] = 0 if coluna in {"estoque", "desconto", "pf_dist", "pf_fabrica", "preco_com_imposto", "preco_sem_imposto"} else ""

    base["cnpj_referencia"] = base["cnpj_referencia"].apply(normalizar_cnpj)
    base["ean"] = base["ean"].apply(normalizar_ean)
    for coluna in ["produto", "distribuidora", "consultor", "uf", "status", "erro"]:
        base[coluna] = base[coluna].apply(_texto)
    base["uf"] = base["uf"].str.upper()
    for coluna in ["estoque", "desconto", "pf_dist", "pf_fabrica", "preco_com_imposto", "preco_sem_imposto"]:
        base[coluna] = base[coluna].apply(converter_numero)
    base["desconto"] = base["desconto"].where(base["desconto"] <= 1, base["desconto"] / 100)
    base["data_atualizacao"] = pd.to_datetime(base["data_atualizacao"], errors="coerce", dayfirst=True)
    return base[COLUNAS_MERCADO].reset_index(drop=True)


def mercado_farma_atual() -> pd.DataFrame:
    return preparar_mercado_farma(carregar_mercado_farma())


def carregar_status_consolidado() -> dict:
    if not MERCADO_FARMA_STATUS.exists():
        return {}
    try:
        import json

        dados = json.loads(MERCADO_FARMA_STATUS.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return dados if isinstance(dados, dict) else {}


def carregar_descontos_adicionais() -> dict:
    dados = carregar_json(DESCONTOS_KEY, {"distribuidoras": {}})
    if not isinstance(dados, dict):
        return {"distribuidoras": {}}
    dados.setdefault("distribuidoras", {})
    return dados


def salvar_descontos_adicionais(dados: dict) -> None:
    if not isinstance(dados, dict):
        dados = {"distribuidoras": {}}
    dados.setdefault("distribuidoras", {})
    salvar_json(DESCONTOS_KEY, dados, "Atualiza descontos adicionais Mercado Farma")


def aplicar_descontos_adicionais(df: pd.DataFrame, configuracao: dict | None = None) -> pd.DataFrame:
    base = preparar_mercado_farma(df)
    if base.empty:
        return base
    dados = configuracao if isinstance(configuracao, dict) else carregar_descontos_adicionais()
    regras = dados.get("distribuidoras", {}) if isinstance(dados, dict) else {}
    if not isinstance(regras, dict):
        return base
    for distribuidora, regra in regras.items():
        if not isinstance(regra, dict):
            continue
        percentual = converter_numero(regra.get("percentual", 0))
        if percentual > 1:
            percentual = percentual / 100
        percentual = max(min(percentual, 1), 0)
        if percentual <= 0:
            continue
        sem_desconto = {normalizar_ean(ean) for ean in regra.get("eans_sem_desconto", []) if normalizar_ean(ean)}
        mask = base["distribuidora"].astype(str).eq(str(distribuidora))
        if sem_desconto:
            mask = mask & ~base["ean"].astype(str).isin(sem_desconto)
        fator = 1 - percentual
        base.loc[mask, "preco_sem_imposto"] = base.loc[mask, "preco_sem_imposto"] * fator
        base.loc[mask, "preco_com_imposto"] = base.loc[mask, "preco_com_imposto"] * fator
        base.loc[mask, "desconto"] = (base.loc[mask, "desconto"] + percentual).clip(upper=1)
    return base


def salvar_mercado_farma(df: pd.DataFrame) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    destino = DATA_DIR / "mercado_farma.xlsx"
    base = df[COLUNAS_MERCADO].copy() if not df.empty else pd.DataFrame(columns=COLUNAS_MERCADO)
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Mercado Farma", index=False)
    salvar_bytes("mercado_farma", destino.read_bytes(), "Atualiza Mercado Farma pelo painel")
    try:
        import streamlit as st

        st.cache_data.clear()
    except Exception:
        pass
    return destino


def _salvar_parcial(df: pd.DataFrame) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    base = preparar_mercado_farma(df)
    with pd.ExcelWriter(JOB_PARTIAL_PATH, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Mercado Farma", index=False)


def _carregar_parcial() -> pd.DataFrame:
    if not JOB_PARTIAL_PATH.exists():
        return pd.DataFrame(columns=COLUNAS_MERCADO)
    try:
        return pd.read_excel(JOB_PARTIAL_PATH, dtype=str, engine="openpyxl")
    except Exception:
        return pd.DataFrame(columns=COLUNAS_MERCADO)


def dataframe_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Mercado Farma", index=False)
    return buffer.getvalue()


def _linha_sem_resultado_excel(base: pd.DataFrame) -> pd.Series:
    status_erro = base["status"].astype(str).str.strip().str.upper().isin({"ERRO", "NAO ENCONTRADO"})
    sem_produto = base["produto"].astype(str).str.strip().eq("")
    sem_distribuidora = base["distribuidora"].astype(str).str.strip().eq("")
    sem_valor = (base["estoque"].fillna(0).astype(float) <= 0) & (base["preco_sem_imposto"].fillna(0).astype(float) <= 0)
    erro_generico = base["erro"].astype(str).str.strip().str.lower().isin({"", "message:", "message"})
    return status_erro & sem_produto & sem_distribuidora & sem_valor & erro_generico


def _tabela_excel_mercado(df: pd.DataFrame, *, incluir_consultor: bool = False) -> pd.DataFrame:
    base = preparar_mercado_farma(df)
    if not base.empty:
        mask_nao_encontrado = _linha_sem_resultado_excel(base)
        base.loc[mask_nao_encontrado, "produto"] = "Produto nao encontrado"
        base.loc[mask_nao_encontrado, "status"] = "NAO ENCONTRADO"
        base.loc[mask_nao_encontrado, "erro"] = "EAN nao encontrado no Mercado Farma"

    colunas = {
        "consultor": "Consultor",
        "uf": "UF",
        "cnpj_referencia": "CNPJ referência",
        "ean": "EAN",
        "produto": "Produto",
        "distribuidora": "Distribuidora",
        "estoque": "Estoque",
        "desconto": "Desconto",
        "pf_dist": "PF Dist.",
        "pf_fabrica": "PF Fábrica",
        "preco_com_imposto": "Preço com imposto",
        "preco_sem_imposto": "Preço sem imposto",
        "data_atualizacao": "Atualizado em",
        "status": "Status",
        "erro": "Erro",
    }
    tabela = base.rename(columns=colunas)
    if not incluir_consultor:
        tabela = tabela.drop(columns=["Consultor"], errors="ignore")
    return tabela


def _ordenar_tabela_excel_mercado(base: pd.DataFrame) -> pd.DataFrame:
    if base.empty:
        return base
    ordenada = base.copy()
    if "Status" in ordenada.columns:
        status = ordenada["Status"].fillna("").astype(str).str.strip().str.upper()
        ordenada["_ordem_status"] = np.select([status.eq("OK"), status.eq("NAO ENCONTRADO")], [0, 1], default=2)
    else:
        ordenada["_ordem_status"] = 0
    ordenacao = [col for col in ["UF", "_ordem_status", "Produto", "EAN", "Distribuidora"] if col in ordenada.columns]
    ordenada = ordenada.sort_values(ordenacao, kind="stable").drop(columns=["_ordem_status"], errors="ignore")
    return ordenada.reset_index(drop=True)


def _nome_aba_uf(valor: object) -> str:
    texto = _texto(valor).upper()
    texto = "".join(char for char in texto if char not in r"[]:*?/\\").strip()
    return (texto or "SEM_UF")[:31]


def _formatar_aba_excel_mercado(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0B5D3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    larguras: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            valor = "" if cell.value is None else str(cell.value)
            larguras[cell.column_letter] = min(max(larguras.get(cell.column_letter, 0), len(valor) + 2), 45)
            cell.border = border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for idx in range(1, ws.max_column + 1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = max(larguras.get(letra, 12), 10)
        cabecalho = ws.cell(row=1, column=idx).value
        if cabecalho == "Desconto":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "0.00%"
        elif cabecalho in {"PF Dist.", "PF Fábrica", "Preço com imposto", "Preço sem imposto"}:
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0.00"
        elif cabecalho == "Estoque":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "#,##0"
        elif cabecalho == "Atualizado em":
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for item in cell:
                    item.number_format = "dd/mm/yyyy hh:mm"


def excel_mercado_farma_por_uf(df: pd.DataFrame, *, incluir_consultor: bool = False) -> bytes:
    base = _ordenar_tabela_excel_mercado(_tabela_excel_mercado(df, incluir_consultor=incluir_consultor))

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if base.empty or "UF" not in base.columns:
            base.to_excel(writer, sheet_name="Mercado Farma", index=False)
            _formatar_aba_excel_mercado(writer.book["Mercado Farma"])
        else:
            abas_usadas: set[str] = set()
            ufs = sorted(uf for uf in base["UF"].dropna().astype(str).str.strip().str.upper().unique().tolist() if uf)
            for uf in ufs:
                aba = _nome_aba_uf(uf)
                while aba in abas_usadas:
                    aba = f"{aba[:28]}_{len(abas_usadas) + 1}"
                abas_usadas.add(aba)
                df_uf = base[base["UF"].astype(str).str.upper().eq(uf)].copy()
                df_uf.to_excel(writer, sheet_name=aba, index=False)
                _formatar_aba_excel_mercado(writer.book[aba])
            if not ufs:
                base.to_excel(writer, sheet_name="SEM_UF", index=False)
                _formatar_aba_excel_mercado(writer.book["SEM_UF"])
    return buffer.getvalue()


def obter_eans_para_consulta(produtos_mix: pd.DataFrame) -> list[str]:
    if produtos_mix is None or produtos_mix.empty:
        return []
    base = padronizar_colunas(produtos_mix.copy())
    coluna = "ean_limpo" if "ean_limpo" in base.columns else "ean" if "ean" in base.columns else base.columns[0]
    valores = base[coluna].dropna().astype(str).map(normalizar_ean)
    tamanhos_validos = {8, 12, 13, 14}
    return sorted({ean for ean in valores.tolist() if len(ean) in tamanhos_validos})


def ufs_por_consultor(clientes: pd.DataFrame) -> dict[str, list[dict[str, str]]]:
    if clientes is None or clientes.empty:
        return {}
    base = clientes.copy()
    for coluna in ["nome_rep", "uf", "cnpj_limpo"]:
        if coluna not in base.columns:
            base[coluna] = ""
    base = base[base["cnpj_limpo"].astype(str).str.strip().ne("")].copy()
    if "cliente_ativo" in base.columns:
        base = base[base["cliente_ativo"].fillna(True)].copy()
    retorno: dict[str, list[dict[str, str]]] = {}
    for (consultor, uf), grupo in base.groupby(["nome_rep", "uf"], dropna=False):
        consultor_txt = _texto(consultor) or "SEM CONSULTOR"
        uf_txt = _texto(uf).upper()
        if uf_txt not in VALID_UFS:
            continue
        cnpj = str(grupo["cnpj_limpo"].dropna().astype(str).iloc[0])
        retorno.setdefault(consultor_txt, []).append({"uf": uf_txt, "cnpj": cnpj})
    for consultor in retorno:
        retorno[consultor] = sorted(retorno[consultor], key=lambda item: item["uf"])
    return retorno


def _clientes_ativos_com_cnpj_por_uf(clientes: pd.DataFrame) -> pd.DataFrame:
    if clientes is None or clientes.empty:
        return pd.DataFrame(columns=["uf", "cnpj_limpo"])
    base = clientes.copy()
    for coluna in ["uf", "cnpj_limpo"]:
        if coluna not in base.columns:
            base[coluna] = ""
    if "cliente_ativo" in base.columns:
        base = base[base["cliente_ativo"].fillna(True)].copy()
    base["uf"] = base["uf"].astype(str).str.strip().str.upper()
    base["cnpj_limpo"] = base["cnpj_limpo"].apply(normalizar_cnpj)
    base = base[base["uf"].isin(VALID_UFS) & base["cnpj_limpo"].str.len().eq(14)].copy()
    if base.empty:
        return pd.DataFrame(columns=["uf", "cnpj_limpo"])
    ordenacao = [col for col in ["uf", "nome_rep", "nome_pdv", "cnpj_limpo"] if col in base.columns]
    return base.sort_values(ordenacao or ["uf", "cnpj_limpo"]).reset_index(drop=True)


def alvos_mercadofarma_por_uf(
    clientes: pd.DataFrame,
    usuario_gd: str,
    senha_gd: str,
) -> list[dict[str, str]]:
    base = _clientes_ativos_com_cnpj_por_uf(clientes)
    if base.empty:
        return []

    alvos: list[dict[str, str]] = []
    for uf, grupo_uf in base.groupby("uf", dropna=False):
        cnpjs: list[str] = []
        for valor in grupo_uf["cnpj_limpo"].dropna().astype(str):
            cnpj = normalizar_cnpj(valor)
            if cnpj and cnpj not in cnpjs:
                cnpjs.append(cnpj)
        if not cnpjs:
            continue
        alvos.append(
            {
                "consultor": "GD",
                "uf": str(uf),
                "cnpj": cnpjs[0],
                "cnpjs_candidatos": cnpjs,
                "usuario": _texto(usuario_gd),
                "senha": _texto(senha_gd),
            }
        )
    return sorted(alvos, key=lambda item: item["uf"])


def _cnpjs_candidatos_alvo(alvo: dict) -> list[str]:
    valores = alvo.get("cnpjs_candidatos", [])
    if isinstance(valores, str):
        brutos = [item.strip() for item in valores.split(",")]
    elif isinstance(valores, (list, tuple, set)):
        brutos = list(valores)
    else:
        brutos = []

    principal = normalizar_cnpj(alvo.get("cnpj"))
    cnpjs: list[str] = []
    if principal:
        cnpjs.append(principal)
    for valor in brutos:
        cnpj = normalizar_cnpj(valor)
        if cnpj and cnpj not in cnpjs:
            cnpjs.append(cnpj)
    return cnpjs


def alvos_unicos_por_uf(
    clientes: pd.DataFrame,
    credenciais: list[dict[str, str]] | None = None,
    *,
    exigir_login: bool | None = None,
) -> list[dict[str, str]]:
    if clientes is None or clientes.empty:
        return []
    credenciais = credenciais or []
    cred_por_consultor = {
        _norm_nome(item.get("consultor")): item
        for item in credenciais
        if _texto(item.get("usuario")) and _texto(item.get("senha"))
    }
    exigir_login = bool(cred_por_consultor) if exigir_login is None else exigir_login

    base = clientes.copy()
    for coluna in ["nome_rep", "uf", "cnpj_limpo"]:
        if coluna not in base.columns:
            base[coluna] = ""
    if "cliente_ativo" in base.columns:
        base = base[base["cliente_ativo"].fillna(True)].copy()
    base["uf"] = base["uf"].astype(str).str.strip().str.upper()
    base = base[base["uf"].isin(VALID_UFS) & base["cnpj_limpo"].astype(str).str.strip().ne("")]
    if base.empty:
        return []

    alvos: list[dict[str, str]] = []
    for uf, grupo_uf in base.sort_values(["uf", "nome_rep", "cnpj_limpo"]).groupby("uf", dropna=False):
        escolhido: dict[str, str] | None = None
        for _, linha in grupo_uf.iterrows():
            consultor = _texto(linha.get("nome_rep")) or "SEM CONSULTOR"
            cred = cred_por_consultor.get(_norm_nome(consultor), {})
            if exigir_login and not cred:
                continue
            escolhido = {
                "consultor": consultor,
                "uf": str(uf),
                "cnpj": normalizar_cnpj(linha.get("cnpj_limpo")),
                "usuario": _texto(cred.get("usuario", "")),
                "senha": _texto(cred.get("senha", "")),
            }
            break
        if escolhido:
            alvos.append(escolhido)
    return alvos


def criar_driver(headless: bool = True):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1366,900")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-background-networking")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    chrome_bin = (
        os.environ.get("CHROME_BIN")
        or shutil.which("google-chrome")
        or shutil.which("google-chrome-stable")
        or shutil.which("chrome")
        or shutil.which("chromium")
        or shutil.which("chromium-browser")
    )
    if chrome_bin:
        options.binary_location = chrome_bin

    erros: list[str] = []
    driver_path = os.environ.get("CHROMEDRIVER_PATH")
    if driver_path:
        try:
            return webdriver.Chrome(service=Service(driver_path), options=options)
        except Exception as exc:
            erros.append(f"CHROMEDRIVER_PATH ({driver_path}): {exc}")

    try:
        return webdriver.Chrome(options=options)
    except Exception as exc:
        erros.append(f"Selenium Manager: {exc}")

    driver_path = shutil.which("chromedriver")
    if driver_path:
        try:
            return webdriver.Chrome(service=Service(driver_path), options=options)
        except Exception as exc:
            erros.append(f"chromedriver do sistema ({driver_path}): {exc}")

    try:
        from webdriver_manager.chrome import ChromeDriverManager

        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    except Exception as exc:
        erros.append(f"webdriver-manager: {exc}")
        resumo = " | ".join(str(erro).splitlines()[0] for erro in erros if erro)
        raise RuntimeError(f"Não consegui abrir o navegador para o Mercado Farma. {resumo}") from exc


def converter_linhas_extrator(linhas: list[dict], consultor: str, uf: str, cnpj: str) -> list[dict]:
    agora = agora_brasilia().strftime("%d/%m/%Y %H:%M:%S")
    convertidas: list[dict] = []
    for item in linhas:
        convertidas.append(
            {
                "consultor": consultor,
                "uf": uf,
                "cnpj_referencia": cnpj,
                "ean": normalizar_ean(item.get("EAN", "")),
                "produto": _texto(item.get("NOME DO PRODUTO", "")),
                "distribuidora": _texto(item.get("DISTRIBUIDORA", "")),
                "estoque": converter_numero(item.get("ESTOQUE", 0)),
                "desconto": converter_numero(item.get("DESCONTO (%)", 0)),
                "pf_dist": converter_numero(item.get("PF DIST. (R$)", 0)),
                "pf_fabrica": converter_numero(item.get("PF FABRICA (R$)", 0)),
                "preco_com_imposto": converter_numero(item.get("PREÇO FINAL (R$)", item.get("PREÃ‡O FINAL (R$)", 0))),
                "preco_sem_imposto": converter_numero(item.get("SEM IMPOSTO (R$)", 0)),
                "data_atualizacao": agora,
                "status": _texto(item.get("STATUS", "OK")) or "OK",
                "erro": _texto(item.get("ERRO", "")),
            }
        )
    return convertidas


def extrair_mercado_farma(
    credenciais: list[dict[str, str]],
    clientes: pd.DataFrame,
    produtos_mix: pd.DataFrame,
    *,
    headless: bool = True,
    limite_eans: int | None = None,
    log_fn: Callable[[str], None] | None = None,
) -> Path:
    eans = obter_eans_para_consulta(produtos_mix)
    if limite_eans:
        eans = eans[: int(limite_eans)]
    if not eans:
        raise RuntimeError("Importe Produtos / Mix para gerar lista de EANs.")

    usuario_gd, senha_gd = _credenciais_gd_da_lista(credenciais)
    if not usuario_gd or not senha_gd:
        gd = carregar_credenciais_mercadofarma(exigir=True)
        usuario_gd = str(gd.get("usuario", ""))
        senha_gd = str(gd.get("senha", ""))

    alvos_unicos = alvos_mercadofarma_por_uf(clientes, usuario_gd, senha_gd)
    if not alvos_unicos:
        raise RuntimeError("Não encontrei CNPJ referência ativo na base de clientes.")

    resultados: list[dict] = []
    for alvo in alvos_unicos:
        resultados.extend(_extrair_alvo(alvo, eans, headless=headless, log_fn=log_fn))

    if not resultados:
        raise RuntimeError("Nenhum preço foi extraído. Verifique logins, senhas e CNPJs.")
    return salvar_mercado_farma(preparar_mercado_farma(pd.DataFrame(resultados)))


def _mensagem_erro_consulta(erro: Exception | str) -> str:
    texto = str(erro or "").strip()
    if texto.lower() in {"", "message:", "message"}:
        return "Nao consegui consultar o EAN no Mercado Farma."
    return texto


def _linha_erro(alvo: dict[str, str], ean: str, erro: Exception | str) -> dict:
    return {
        "consultor": _texto(alvo.get("consultor")),
        "uf": _texto(alvo.get("uf")).upper(),
        "cnpj_referencia": normalizar_cnpj(alvo.get("cnpj")),
        "ean": normalizar_ean(ean),
        "produto": "",
        "distribuidora": "",
        "estoque": 0,
        "desconto": 0,
        "pf_dist": 0,
        "pf_fabrica": 0,
        "preco_com_imposto": 0,
        "preco_sem_imposto": 0,
        "data_atualizacao": agora_brasilia().strftime("%d/%m/%Y %H:%M:%S"),
        "status": "ERRO",
        "erro": _mensagem_erro_consulta(erro),
    }


def _salvar_debug_driver(driver, debug_dir: Path | str | None, nome: str) -> None:
    if driver is None or debug_dir is None:
        return
    pasta = Path(debug_dir)
    pasta.mkdir(parents=True, exist_ok=True)
    try:
        driver.save_screenshot(str(pasta / f"{nome}.png"))
    except Exception:
        pass
    try:
        (pasta / "debug_html.html").write_text(driver.page_source or "", encoding="utf-8")
    except Exception:
        pass


def _extrair_alvo(
    alvo: dict[str, str],
    eans: list[str],
    *,
    headless: bool,
    log_fn: Callable[[str], None] | None = None,
    start_index: int = 0,
    estado: dict | None = None,
    resultados: list[dict] | None = None,
    debug_dir: Path | str | None = None,
) -> list[dict]:
    saida = resultados if resultados is not None else []
    driver = None
    consultor = _texto(alvo.get("consultor"))
    uf = _texto(alvo.get("uf")).upper()
    cnpj = normalizar_cnpj(alvo.get("cnpj"))
    cnpjs_candidatos = _cnpjs_candidatos_alvo(alvo)
    usuario = _texto(alvo.get("usuario"))
    senha = _texto(alvo.get("senha"))
    etapa = "abrir_navegador"
    try:
        if not cnpjs_candidatos:
            raise RuntimeError(f"UF {uf}: nenhum CNPJ candidato valido para Mercado Farma.")
        if callable(log_fn):
            log_fn(
                f"UF {uf}: CNPJ referencia {cnpj} | "
                f"{len(cnpjs_candidatos)} candidato(s) | usuario {mascarar_usuario(usuario)}"
            )
            log_fn(f"UF {uf}: etapa login - abrindo Mercado Farma")
        driver = criar_driver(headless=headless)
        etapa = "login"
        login_mercadofarma(driver, usuario, senha, log_fn=log_fn)
        etapa = "selecao_cnpj"
        erros_cnpj: list[str] = []
        for posicao, candidato in enumerate(cnpjs_candidatos, start=1):
            cnpj = candidato
            if posicao > 1:
                if callable(log_fn):
                    log_fn(f"UF {uf}: resetando tela de selecao antes do proximo CNPJ.")
                driver.get("https://www.mercadofarma.com.br/selecionar-loja")
            if estado is not None:
                estado["current_cnpj"] = cnpj
            if callable(log_fn):
                log_fn(f"UF {uf}: etapa selecao CNPJ {posicao}/{len(cnpjs_candidatos)} - {cnpj}")
            try:
                selecionar_cnpj_catalogo(driver, cnpj, log_fn=log_fn)
                alvo["cnpj"] = cnpj
                break
            except Exception as exc:
                erros_cnpj.append(f"{cnpj}: {_mensagem_erro_consulta(exc)}")
                if callable(log_fn):
                    log_fn(f"UF {uf}: CNPJ {cnpj} indisponivel no Mercado Farma. Tentando proximo.")
        else:
            detalhes = " | ".join(erros_cnpj[-5:])
            raise RuntimeError(f"UF {uf}: nenhum CNPJ candidato abriu o Mercado Farma. {detalhes}")

        etapa = "catalogo"
        if callable(log_fn):
            log_fn(f"UF {uf}: etapa catalogo - Catalogo A a Z carregado com CNPJ {cnpj}")
        etapa = "extracao"
        for idx in range(start_index, len(eans)):
            ean = eans[idx]
            if estado is not None and estado.get("cancelar"):
                raise RuntimeError("Extração cancelada pelo usuário.")
            job_id = str(estado.get("job_id", "")) if estado is not None else ""
            cancel_event = _CANCEL_FLAGS.get(job_id)
            if cancel_event is not None and cancel_event.is_set():
                raise RuntimeError("Extração cancelada pelo usuário.")
            if estado is not None:
                estado["ean_index"] = idx
                estado["current_ean"] = ean
                estado["processados"] = int(estado.get("target_index", 0)) * len(eans) + idx
            if callable(log_fn):
                log_fn(f"UF {uf}: etapa extracao - consultando {idx + 1}/{len(eans)} - {ean}")
            try:
                linhas = processar_ean_catalogo(driver, ean)
                saida.extend(converter_linhas_extrator(linhas, consultor, uf, cnpj))
            except Exception as exc:
                if debug_dir is not None and idx == start_index:
                    _salvar_debug_driver(driver, debug_dir, "erro_busca")
                saida.append(_linha_erro(alvo, ean, exc))
            if estado is not None:
                estado["ean_index"] = idx + 1
                estado["processados"] = int(estado.get("target_index", 0)) * len(eans) + idx + 1
            if estado is not None and (idx + 1) % FLUSH_EVERY_EANS == 0:
                _salvar_parcial(pd.DataFrame(saida))
                _salvar_estado_extracao(estado)
        return saida
    except Exception as exc:
        nome_debug = {
            "login": "erro_login",
            "selecao_cnpj": "erro_selecao_cnpj",
            "catalogo": "erro_catalogo",
            "extracao": "erro_extracao",
            "abrir_navegador": "erro_navegador",
        }.get(etapa, "erro_extracao")
        _salvar_debug_driver(driver, debug_dir, nome_debug)
        raise RuntimeError(f"{etapa}: {exc}") from exc
    finally:
        if driver is not None:
            driver.quit()


def iniciar_extracao_background(
    credenciais: list[dict[str, str]],
    clientes: pd.DataFrame,
    produtos_mix: pd.DataFrame,
    *,
    headless: bool = True,
    limite_eans: int | None = None,
    retomar: bool = False,
    ufs: list[str] | None = None,
) -> dict:
    atual = carregar_estado_extracao()
    if atual.get("status") == "rodando" and atual.get("thread_alive"):
        return atual

    eans = obter_eans_para_consulta(produtos_mix)
    if limite_eans:
        eans = eans[: int(limite_eans)]
    usuario_gd, senha_gd = _credenciais_gd_da_lista(credenciais)
    if not usuario_gd or not senha_gd:
        gd = carregar_credenciais_mercadofarma(exigir=True)
        usuario_gd = str(gd.get("usuario", ""))
        senha_gd = str(gd.get("senha", ""))
    alvos = alvos_mercadofarma_por_uf(clientes, usuario_gd, senha_gd)
    if ufs:
        ufs_set = {str(uf).strip().upper() for uf in ufs if str(uf).strip()}
        alvos = [alvo for alvo in alvos if str(alvo.get("uf", "")).upper() in ufs_set]
    if not eans:
        raise RuntimeError("Importe Produtos / Mix para gerar lista de EANs.")
    if not alvos:
        raise RuntimeError("Não encontrei UF válida com CNPJ referência ativo.")

    job_id = uuid4().hex
    start_target = int(atual.get("target_index", 0) or 0) if retomar else 0
    start_ean = int(atual.get("ean_index", 0) or 0) if retomar else 0
    if not retomar and JOB_PARTIAL_PATH.exists():
        JOB_PARTIAL_PATH.unlink()

    estado = _estado_padrao()
    estado.update(
        {
            "job_id": job_id,
            "status": "rodando",
            "inicio": agora_brasilia().isoformat(),
            "total_eans": len(eans),
            "total_alvos": len(alvos),
            "total_passos": len(eans) * len(alvos),
            "target_index": start_target,
            "ean_index": start_ean,
            "alvos": [{k: alvo.get(k, "") for k in ["consultor", "uf", "cnpj"]} for alvo in alvos],
            "cancelar": False,
        }
    )
    _log_estado(estado, "Extração iniciada em segundo plano.")
    _salvar_estado_extracao(estado)

    _CANCEL_FLAGS[job_id] = threading.Event()
    thread = threading.Thread(
        target=_worker_extracao,
        args=(job_id, alvos, eans, headless, retomar, start_target, start_ean),
        daemon=True,
    )
    _THREADS[job_id] = thread
    thread.start()
    return carregar_estado_extracao()


def cancelar_extracao_background() -> dict:
    estado = carregar_estado_extracao()
    estado["cancelar"] = True
    estado["mensagem"] = "Cancelamento solicitado. A extração vai parar ao finalizar o item atual."
    job_id = str(estado.get("job_id", ""))
    if job_id in _CANCEL_FLAGS:
        _CANCEL_FLAGS[job_id].set()
    _salvar_estado_extracao(estado)
    return estado


def _worker_extracao(
    job_id: str,
    alvos: list[dict[str, str]],
    eans: list[str],
    headless: bool,
    retomar: bool,
    start_target: int,
    start_ean: int,
) -> None:
    estado = carregar_estado_extracao()
    resultados: list[dict] = []
    if retomar:
        parcial = _carregar_parcial()
        resultados = [] if parcial.empty else preparar_mercado_farma(parcial).to_dict("records")
    try:
        for target_idx in range(start_target, len(alvos)):
            alvo = alvos[target_idx]
            estado.update(
                {
                    "job_id": job_id,
                    "status": "rodando",
                    "target_index": target_idx,
                    "ean_index": start_ean if target_idx == start_target else 0,
                    "current_uf": alvo.get("uf", ""),
                    "current_consultor": alvo.get("consultor", ""),
                    "current_cnpj": alvo.get("cnpj", ""),
                }
            )
            _log_estado(estado, f"UF {alvo.get('uf')}: iniciando com CNPJ {alvo.get('cnpj')}.")
            _salvar_estado_extracao(estado)
            _extrair_alvo(
                alvo,
                eans,
                headless=headless,
                start_index=start_ean if target_idx == start_target else 0,
                estado=estado,
                resultados=resultados,
            )
            start_ean = 0
            estado["target_index"] = target_idx + 1
            estado["ean_index"] = 0
            estado["processados"] = min((target_idx + 1) * len(eans), int(estado.get("total_passos", 0) or 0))
            _salvar_parcial(pd.DataFrame(resultados))
            _salvar_estado_extracao(estado)
            if estado.get("cancelar"):
                raise RuntimeError("Extração cancelada pelo usuário.")

        destino = salvar_mercado_farma(preparar_mercado_farma(pd.DataFrame(resultados)))
        estado["status"] = "concluido"
        estado["fim"] = agora_brasilia().isoformat()
        estado["processados"] = int(estado.get("total_passos", 0) or 0)
        estado["current_ean"] = ""
        estado["erro"] = ""
        _log_estado(estado, f"Extração concluída e salva em {destino.name}.")
        _salvar_estado_extracao(estado)
    except Exception as exc:
        if resultados:
            _salvar_parcial(pd.DataFrame(resultados))
        estado["status"] = "cancelado" if "cancelada" in str(exc).lower() else "erro"
        estado["fim"] = agora_brasilia().isoformat()
        estado["erro"] = str(exc)
        _log_estado(estado, f"Extração interrompida: {exc}")
        _salvar_estado_extracao(estado)
    finally:
        _CANCEL_FLAGS.pop(job_id, None)


def melhor_preco_por_ean(df: pd.DataFrame) -> pd.DataFrame:
    base = preparar_mercado_farma(df)
    if base.empty:
        return base
    validos = base[(base["estoque"] > 0) & (base["preco_sem_imposto"] > 0)].copy()
    if validos.empty:
        return pd.DataFrame(columns=COLUNAS_MERCADO)
    return validos.sort_values(["uf", "ean", "preco_sem_imposto", "estoque"], ascending=[True, True, True, False]).drop_duplicates(["uf", "ean"])


def formatar_tabela_mercado(df: pd.DataFrame) -> pd.DataFrame:
    base = preparar_mercado_farma(df)
    colunas = {
        "consultor": "Consultor",
        "uf": "UF",
        "cnpj_referencia": "CNPJ referência",
        "ean": "EAN",
        "produto": "Produto",
        "distribuidora": "Distribuidora",
        "estoque": "Estoque",
        "desconto": "Desconto",
        "pf_dist": "PF Dist.",
        "pf_fabrica": "PF Fábrica",
        "preco_com_imposto": "Preço com imposto",
        "preco_sem_imposto": "Preço sem imposto",
        "data_atualizacao": "Atualizado em",
        "status": "Status",
        "erro": "Erro",
    }
    base = base.rename(columns=colunas)
    for coluna in ["PF Dist.", "PF Fábrica", "Preço com imposto", "Preço sem imposto"]:
        if coluna in base.columns:
            base[coluna] = base[coluna].apply(formatar_moeda)
    if "Desconto" in base.columns:
        base["Desconto"] = base["Desconto"].apply(lambda valor: f"{float(valor or 0) * 100:,.2f}%".replace(",", "X").replace(".", ",").replace("X", "."))
    if "Atualizado em" in base.columns:
        base["Atualizado em"] = pd.to_datetime(base["Atualizado em"], errors="coerce", dayfirst=True).dt.strftime("%d/%m/%Y %H:%M")
        base["Atualizado em"] = base["Atualizado em"].fillna("-")
    return base
